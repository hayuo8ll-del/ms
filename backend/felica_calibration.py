"""生成したボトルネック計画を、現場の実計画(FeliCa)と製番単位で突き合わせて
精度を測り、工程オフセット/A勤割合を較正する。

FeliCa は `YYYYMM_CTA{1,2}` シートに、製番(Plan列)ごとの Line-In(投入=開始)/
Completion(完成≒MIL) の日次台数を持つ。台帳から作った計画の MIL完成日・ANT投入日と
突き合わせ、稼働日インデックス差の平均絶対誤差(MAE)で精度を評価する。

較正結果は推奨値として返すだけで、config は自動書き換えしない(手動反映)。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import BinaryIO

from openpyxl import load_workbook

from bottleneck_planner import (
    DemandItem,
    StageFlowConfig,
    plan_bottleneck,
    working_days_in_range,
)


@dataclass
class FelicaLot:
    product: str
    lot: float
    line_in_first: date | None
    completion_last: date | None
    line_in_daily: dict[date, float] = field(default_factory=dict)
    completion_daily: dict[date, float] = field(default_factory=dict)


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_felica_plan(file_obj: BinaryIO) -> dict[str, FelicaLot]:
    """FeliCa実計画を製番別に読む(全 `*_CTA*` シートをまたいで合算)。"""
    wb = load_workbook(file_obj, data_only=True)
    lots: dict[str, FelicaLot] = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        # 日付ヘッダー(行3, col9以降)
        date_cols: dict[int, date] = {}
        for c in range(9, ws.max_column + 1):
            d = _as_date(ws.cell(row=3, column=c).value)
            if d:
                date_cols[c] = d
        if not date_cols:
            continue

        r = 5
        while r <= ws.max_row:
            role = ws.cell(row=r, column=8).value
            seiban = ws.cell(row=r, column=4).value
            if role == "Line-In" and seiban:
                key = str(seiban).strip()
                product = ws.cell(row=r, column=3).value
                lot_size = ws.cell(row=r, column=7).value
                entry = lots.get(key)
                if entry is None:
                    entry = FelicaLot(
                        product=str(product).strip() if product else "",
                        lot=float(lot_size) if isinstance(lot_size, (int, float)) else 0.0,
                        line_in_first=None,
                        completion_last=None,
                    )
                    lots[key] = entry
                # Line-In(この行) と Completion(次の行) の日次を拾う
                for c, d in date_cols.items():
                    li = ws.cell(row=r, column=c).value
                    if isinstance(li, (int, float)) and li:
                        entry.line_in_daily[d] = entry.line_in_daily.get(d, 0.0) + float(li)
                    co = ws.cell(row=r + 1, column=c).value
                    if isinstance(co, (int, float)) and co:
                        entry.completion_daily[d] = entry.completion_daily.get(d, 0.0) + float(co)
                r += 2
            else:
                r += 1

    for entry in lots.values():
        if entry.line_in_daily:
            entry.line_in_first = min(entry.line_in_daily)
        if entry.completion_daily:
            entry.completion_last = max(entry.completion_daily)
    return lots


@dataclass
class ComparisonReport:
    matched: int
    completion_mae: float  # 完成日の平均絶対誤差(稼働日)
    completion_bias: float  # 平均符号付き差(our - felica; 正=遅い)
    start_mae: float  # 投入日の平均絶対誤差(稼働日)
    start_bias: float
    completion_daily_mae: float  # 日次形状(MIL vs Completion; 予実の重複窓に限定)
    line_in_daily_mae: float  # 日次形状(ANT vs Line-In; 予実の重複窓に限定)
    # 機種別 日次形状MAE(重複窓): {呼称: {"completion_mae", "line_in_mae", "days"}}
    daily_shape_by_product: dict[str, dict] = field(default_factory=dict)


def _wd_index_map(working_days: list[date]) -> dict[date, int]:
    return {d: i for i, d in enumerate(working_days)}


def _nearest_index(d: date, wd_index: dict[date, int], working_days: list[date]) -> int | None:
    if d in wd_index:
        return wd_index[d]
    # 稼働日でない日(土日)は直近の稼働日インデックスへ丸める
    if not working_days:
        return None
    return min(range(len(working_days)), key=lambda i: abs((working_days[i] - d).days))


def _windowed_daily_mae(a: dict[date, float], b: dict[date, float]) -> tuple[float, int]:
    """2つの日次系列を、両者の活動日レンジの**重複窓**に限定して MAE を出す。

    計画窓とFeliCaの月レンジがずれると、非重複の裾(片側だけ台数がある日)が誤差を水増し
    するため、`lo=max(min(a),min(b))`〜`hi=min(max(a),max(b))` の範囲に絞る。窓内で
    どちらかに台数がある日だけを母数にする(両方ゼロの日は比較対象にしない)。
    戻り値 = (MAE, 対象日数)。重複が無ければ (0, 0)。
    """
    if not a or not b:
        return 0.0, 0
    lo = max(min(a), min(b))
    hi = min(max(a), max(b))
    if lo > hi:
        return 0.0, 0
    days = [d for d in (set(a) | set(b)) if lo <= d <= hi]
    if not days:
        return 0.0, 0
    return sum(abs(a.get(d, 0.0) - b.get(d, 0.0)) for d in days) / len(days), len(days)


def compare_plans(
    result,
    felica: dict[str, FelicaLot],
    working_days: list[date],
    aliases: dict[str, str] | None = None,
) -> ComparisonReport:
    """生成計画と FeliCa 実計画を製番単位で突き合わせて誤差を出す。

    `aliases` を渡すと、FeliCa の Item Desc(RCコード)を呼称へ解決して機種別の日次形状MAEも出す
    (our の機種名は `stage_allocation` の product=呼称)。省略時は機種別内訳は空。
    """
    from thm_ledger_import import resolve_product

    wd_index = _wd_index_map(working_days)

    # our: 製番 -> MIL完成日 / ANT最早投入日、および工程×機種の日次
    our_completion = {lot.order_id: lot.completion_day for lot in result.mil_lots if lot.order_id}
    our_ant_start: dict[str, date] = {}
    our_mil_daily: dict[date, float] = {}
    our_ant_daily: dict[date, float] = {}
    our_mil_by_product: dict[str, dict[date, float]] = {}
    our_ant_by_product: dict[str, dict[date, float]] = {}
    for c in result.stage_allocation:
        if c.stage_id == "ANT":
            our_ant_daily[c.day] = our_ant_daily.get(c.day, 0.0) + c.quantity
            series = our_ant_by_product.setdefault(c.product, {})
            series[c.day] = series.get(c.day, 0.0) + c.quantity
            if c.order_id and (c.order_id not in our_ant_start or c.day < our_ant_start[c.order_id]):
                our_ant_start[c.order_id] = c.day
        elif c.stage_id == "MIL":
            our_mil_daily[c.day] = our_mil_daily.get(c.day, 0.0) + c.quantity
            series = our_mil_by_product.setdefault(c.product, {})
            series[c.day] = series.get(c.day, 0.0) + c.quantity

    comp_diffs: list[int] = []
    start_diffs: list[int] = []
    for seiban, fl in felica.items():
        if seiban in our_completion and fl.completion_last is not None:
            oi = _nearest_index(our_completion[seiban], wd_index, working_days)
            fi = _nearest_index(fl.completion_last, wd_index, working_days)
            if oi is not None and fi is not None:
                comp_diffs.append(oi - fi)
        if seiban in our_ant_start and fl.line_in_first is not None:
            oi = _nearest_index(our_ant_start[seiban], wd_index, working_days)
            fi = _nearest_index(fl.line_in_first, wd_index, working_days)
            if oi is not None and fi is not None:
                start_diffs.append(oi - fi)

    # FeliCa 日次形状(ライン計 + 機種別)
    fel_comp_daily: dict[date, float] = {}
    fel_li_daily: dict[date, float] = {}
    fel_comp_by_product: dict[str, dict[date, float]] = {}
    fel_li_by_product: dict[str, dict[date, float]] = {}
    for fl in felica.values():
        product = resolve_product(fl.product, aliases) if aliases is not None else None
        for d, q in fl.completion_daily.items():
            fel_comp_daily[d] = fel_comp_daily.get(d, 0.0) + q
            if product:
                s = fel_comp_by_product.setdefault(product, {})
                s[d] = s.get(d, 0.0) + q
        for d, q in fl.line_in_daily.items():
            fel_li_daily[d] = fel_li_daily.get(d, 0.0) + q
            if product:
                s = fel_li_by_product.setdefault(product, {})
                s[d] = s.get(d, 0.0) + q

    # 機種別 日次形状MAE(重複窓)
    daily_shape_by_product: dict[str, dict] = {}
    if aliases is not None:
        for product in set(our_mil_by_product) | set(fel_comp_by_product) | set(our_ant_by_product) | set(fel_li_by_product):
            comp_mae, comp_n = _windowed_daily_mae(
                our_mil_by_product.get(product, {}), fel_comp_by_product.get(product, {})
            )
            li_mae, li_n = _windowed_daily_mae(
                our_ant_by_product.get(product, {}), fel_li_by_product.get(product, {})
            )
            if comp_n == 0 and li_n == 0:
                continue  # 予実の重複が無い機種は出さない
            daily_shape_by_product[product] = {
                "completion_mae": round(comp_mae, 0),
                "line_in_mae": round(li_mae, 0),
                "days": max(comp_n, li_n),
            }

    def mae(xs: list[int]) -> float:
        return sum(abs(x) for x in xs) / len(xs) if xs else 0.0

    def bias(xs: list[int]) -> float:
        return sum(xs) / len(xs) if xs else 0.0

    return ComparisonReport(
        matched=len(comp_diffs),
        completion_mae=round(mae(comp_diffs), 2),
        completion_bias=round(bias(comp_diffs), 2),
        start_mae=round(mae(start_diffs), 2),
        start_bias=round(bias(start_diffs), 2),
        completion_daily_mae=round(_windowed_daily_mae(our_mil_daily, fel_comp_daily)[0], 0),
        line_in_daily_mae=round(_windowed_daily_mae(our_ant_daily, fel_li_daily)[0], 0),
        daily_shape_by_product=daily_shape_by_product,
    )


def _median(xs: list[int]) -> int:
    s = sorted(xs)
    n = len(s)
    return int(round((s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2)))


def derive_stage_offsets(
    felica: dict[str, FelicaLot],
    working_days: list[date],
    current_flows: list[StageFlowConfig],
    aliases: dict[str, str] | None = None,
    min_samples: int = 2,
) -> dict[str, dict[str, int]]:
    """FeliCaの実投入→完成の稼働日スパンから、機種別の工程オフセットを実測で導出する。

    製番ごとの (Completion − Line-In) を稼働日で測り、機種ごとに中央値=実リードLを求める。
    現状オフセット(例 ANT-2/HAL0/MIL+1, 総スパン3)の各工程比率を保ったまま L で割り付ける
    (ANTは上流に -round(L×上流比率)、MILは下流に +round(L×下流比率)、HALは0)。
    FeliCaのItem Desc(RC-コード)は呼称へ解決して集計する。
    `stageFlows[].leadOffsetByProduct` に入れられる {stage_id: {機種: off}} を返す。
    """
    from thm_ledger_import import resolve_product

    wd_index = _wd_index_map(working_days)

    spans: dict[str, list[int]] = {}
    for fl in felica.values():
        product = resolve_product(fl.product, aliases)
        if fl.line_in_first is None or fl.completion_last is None or not product:
            continue
        li = _nearest_index(fl.line_in_first, wd_index, working_days)
        co = _nearest_index(fl.completion_last, wd_index, working_days)
        if li is None or co is None:
            continue
        spans.setdefault(product, []).append(max(co - li, 0))
    lead_by_product = {p: _median(xs) for p, xs in spans.items() if len(xs) >= min_samples}

    # 現状オフセットの総スパンと各工程比率
    ant_off = min((f.lead_offset_days for f in current_flows), default=-2)
    mil_off = max((f.lead_offset_days for f in current_flows), default=1)
    total_span = mil_off - ant_off or 1

    result: dict[str, dict[str, int]] = {}
    for flow in current_flows:
        if flow.lead_offset_days == 0:  # HAL(基準)は0固定
            continue
        frac = flow.lead_offset_days / total_span  # 上流は負、下流は正
        result[flow.stage_id] = {
            product: int(round(lead * frac)) for product, lead in lead_by_product.items()
        }
    return result


@dataclass
class CalibrationResult:
    current: ComparisonReport
    recommended: ComparisonReport
    recommended_offsets: dict[str, int]  # {ANT, TAL, HAL, MIL}
    recommended_a_shift_fraction: float


def _with_offsets(stage_flows: list[StageFlowConfig], offsets: dict[str, int]) -> list[StageFlowConfig]:
    return [
        StageFlowConfig(f.stage_id, offsets.get(f.stage_id, f.lead_offset_days), f.daily_capacity, f.input_unit)
        for f in stage_flows
    ]


def calibrate(
    demands: list[DemandItem],
    working_days: list[date],
    shift_capacities: dict[str, float],
    plan_kwargs: dict,
    felica: dict[str, FelicaLot],
    ant_range=(-3, -2, -1),
    tal_range=(-2, -1),
    mil_range=(0, 1, 2),
    a_fractions=(0.4, 0.5, 0.6),
) -> CalibrationResult:
    """オフセット/A勤割合をグリッド探索し、完成日+投入日MAE 最小の推奨値を返す。

    HAL は 0 固定(ボトルネック基準)。設備停止は照合では無効化する。
    plan_kwargs は base の plan_bottleneck キーワード引数(stage_flows/equipment_stops 等)。
    shift_capacities はライン日次能力({モード: 台/日})。
    """
    base_flows = plan_kwargs["stage_flows"]
    # stage_flows / equipment_stops / a_shift_fraction は明示指定するので除外
    kw = {k: v for k, v in plan_kwargs.items() if k not in ("stage_flows", "equipment_stops", "a_shift_fraction")}

    def run(flows, a_frac) -> ComparisonReport:
        res = plan_bottleneck(
            demands, working_days, shift_capacities, stage_flows=flows,
            equipment_stops=None, a_shift_fraction=a_frac, **kw
        )
        return compare_plans(res, felica, working_days)

    current = run(base_flows, plan_kwargs.get("a_shift_fraction", 0.5))

    best: tuple[float, dict, float, ComparisonReport] | None = None
    for ant in ant_range:
        for tal in tal_range:
            for mil in mil_range:
                offsets = {"ANT": ant, "TAL": tal, "HAL": 0, "MIL": mil}
                flows = _with_offsets(base_flows, offsets)
                for a_frac in a_fractions:
                    rep = run(flows, a_frac)
                    score = rep.completion_mae + rep.start_mae
                    if best is None or score < best[0]:
                        best = (score, offsets, a_frac, rep)

    return CalibrationResult(
        current=current,
        recommended=best[3],
        recommended_offsets=best[1],
        recommended_a_shift_fraction=best[2],
    )
