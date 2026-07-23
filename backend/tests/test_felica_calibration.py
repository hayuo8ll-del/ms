import io
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402

from bottleneck_planner import DemandItem, StageFlowConfig, plan_bottleneck, working_days_in_range  # noqa: E402
from felica_calibration import (  # noqa: E402
    calibrate,
    compare_plans,
    derive_stage_offsets,
    parse_felica_nonworking_days,
    parse_felica_plan,
)

CAPS = {"16h": 90000, "22h": 120000}


def _felica_workbook(rows):
    """FeliCa形式(行3ヘッダ, 製番=col4, Line-In/Completion=col8 の2行ペア)のブックを作る。

    rows: list of (seiban, product, lot, {date: line_in}, {date: completion})
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "202607_CTA1"
    dates = sorted({d for _s, _p, _l, li, co in rows for d in list(li) + list(co)})
    # 行3: A..H はラベル、col9以降が日付
    for c, label in enumerate(["Planner", "WS", "Item Desc", "Plan", "Item No", "Before", "Lot", "Date"], 1):
        ws.cell(row=3, column=c, value=label)
    date_col = {d: 9 + i for i, d in enumerate(dates)}
    for d, c in date_col.items():
        ws.cell(row=3, column=c, value=d)
    r = 5
    for seiban, product, lot, li, co in rows:
        ws.cell(row=r, column=3, value=product)
        ws.cell(row=r, column=4, value=seiban)
        ws.cell(row=r, column=7, value=lot)
        ws.cell(row=r, column=8, value="Line-In")
        ws.cell(row=r + 1, column=8, value="Completion")
        for d, q in li.items():
            ws.cell(row=r, column=date_col[d], value=q)
        for d, q in co.items():
            ws.cell(row=r + 1, column=date_col[d], value=q)
        r += 2
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_parse_felica_plan_extracts_line_in_and_completion():
    buf = _felica_workbook([
        ("S1", "RC-S982F/5", 40000, {date(2026, 7, 23): 40000}, {date(2026, 7, 28): 40000}),
    ])
    felica = parse_felica_plan(buf)
    assert set(felica) == {"S1"}
    lot = felica["S1"]
    assert lot.line_in_first == date(2026, 7, 23)
    assert lot.completion_last == date(2026, 7, 28)
    assert lot.lot == 40000


def test_parse_felica_excludes_same_day_carryover():
    """同一日にLine-InとCompletionの両方に台数がある日=先週までの計画台数として除外する。"""
    buf = _felica_workbook([
        # 丸ごと同日 → 除外され line_in_first/completion_last が None
        ("C1", "RC-S103/JW16", 57600, {date(2026, 7, 16): 57600}, {date(2026, 7, 16): 57600}),
        # 一部同日(7/16は除外、7/21の完成は残る)
        ("C2", "RC-S127/HCB5", 100000, {date(2026, 7, 16): 100000},
         {date(2026, 7, 16): 50000, date(2026, 7, 21): 50000}),
        # 別日(除外なし)
        ("N1", "RC-SA02F/5", 90000, {date(2026, 7, 16): 90000}, {date(2026, 7, 23): 90000}),
    ])
    felica = parse_felica_plan(buf)

    # C1: 丸ごとcarryover → 投入日/完成日ともに無し
    assert felica["C1"].line_in_first is None
    assert felica["C1"].completion_last is None
    # C2: 7/16は両系列から除去、7/21の完成のみ残る
    assert date(2026, 7, 16) not in felica["C2"].line_in_daily
    assert date(2026, 7, 16) not in felica["C2"].completion_daily
    assert felica["C2"].completion_daily == {date(2026, 7, 21): 50000}
    # N1: 別日なのでそのまま
    assert felica["N1"].line_in_first == date(2026, 7, 16)
    assert felica["N1"].completion_last == date(2026, 7, 23)


def test_compare_plans_reports_completion_day_diff():
    days = working_days_in_range(date(2026, 7, 1), date(2026, 7, 31))
    flows = [StageFlowConfig("ANT", -1), StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 1)]
    demands = [DemandItem("さそり金融", 90000, date(2026, 7, 31), order_id="S1")]
    caps = {"16h": {"さそり金融": 90000}, "22h": {"さそり金融": 120000}}
    result = plan_bottleneck(demands, days, CAPS, stage_flows=flows, product_caps_by_mode=caps)
    our_comp = {lot.order_id: lot.completion_day for lot in result.mil_lots}["S1"]

    # FeliCaの完成日を our より2稼働日後ろに置く → 完成日差 = -2 (ourが早い)
    fi = days.index(our_comp)
    felica_comp = days[fi + 2]
    buf = _felica_workbook([("S1", "RC-SA02F/5", 90000, {days[0]: 90000}, {felica_comp: 90000})])
    felica = parse_felica_plan(buf)

    rep = compare_plans(result, felica, days)
    assert rep.matched == 1
    assert rep.completion_mae == 2.0
    assert rep.completion_bias == -2.0  # our - felica = 早い


def test_calibrate_picks_offsets_reducing_error():
    days = working_days_in_range(date(2026, 7, 1), date(2026, 8, 15))
    base_flows = [StageFlowConfig("ANT", -3), StageFlowConfig("TAL", -2),
                  StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 2)]
    caps = {"16h": {"X": 30000}, "22h": {"X": 45000}}
    demands = [DemandItem("X", 60000, date(2026, 8, 15), order_id="S1")]
    plan_kwargs = dict(stage_flows=base_flows, product_caps_by_mode=caps, a_shift_only_switch=False,
                       a_shift_fraction=0.5)

    # 真値: ANT=-1, MIL=0 で計画した完成日/投入日を FeliCa とする
    truth_flows = [StageFlowConfig("ANT", -1), StageFlowConfig("TAL", -1),
                   StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 0)]
    truth = plan_bottleneck(demands, days, CAPS, stage_flows=truth_flows, product_caps_by_mode=caps)
    comp = {lot.order_id: lot.completion_day for lot in truth.mil_lots}["S1"]
    ant_start = min(c.day for c in truth.stage_allocation if c.stage_id == "ANT" and c.order_id == "S1")
    buf = _felica_workbook([("S1", "RC-S100/5", 60000, {ant_start: 60000}, {comp: 60000})])
    felica = parse_felica_plan(buf)

    cal = calibrate(demands, days, CAPS, plan_kwargs, felica)
    # 較正後の誤差は現状以下
    assert cal.recommended.completion_mae + cal.recommended.start_mae <= cal.current.completion_mae + cal.current.start_mae
    # 真値(MIL=0)に近いオフセットが選ばれる
    assert cal.recommended_offsets["MIL"] == 0


def test_daily_shape_mae_limited_to_overlap_window():
    """予実の重複窓外の裾は日次形状MAEに含めない(union版より小さくなる)。"""
    days = working_days_in_range(date(2026, 7, 1), date(2026, 8, 31))
    flows = [StageFlowConfig("ANT", -1), StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 1)]
    caps = {"16h": {"さそり金融": 90000}, "22h": {"さそり金融": 120000}}
    demands = [DemandItem("さそり金融", 90000, date(2026, 7, 31), order_id="S1")]
    result = plan_bottleneck(demands, days, CAPS, stage_flows=flows, product_caps_by_mode=caps)

    our_comp = {lot.order_id: lot.completion_day for lot in result.mil_lots}["S1"]
    ci = days.index(our_comp)
    # FeliCa: our と重なる完成日 + 遠く離れた裾(非重複)を持たせる
    far = days[-1]  # 8月末、our MIL には無い日
    buf = _felica_workbook([
        ("S1", "RC-SA02F/5", 90000, {days[0]: 90000}, {days[ci]: 60000, far: 30000}),
    ])
    felica = parse_felica_plan(buf)

    rep = compare_plans(result, felica, days)
    # 非重複の裾(far=30000)は窓外なので除外され、重複窓の差(完成日の 90000 vs 60000=30000)が主
    assert rep.completion_daily_mae <= 30000


def test_daily_shape_by_product_breakdown():
    """aliases を渡すと機種別の日次形状MAEが出る(一致機種=0, ズレ機種=正)。"""
    days = working_days_in_range(date(2026, 7, 1), date(2026, 7, 31))
    flows = [StageFlowConfig("ANT", 0), StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 0)]
    caps = {
        "16h": {"さそり金融": 90000, "MOT2": 90000},
        "22h": {"さそり金融": 120000, "MOT2": 120000},
    }
    demands = [
        DemandItem("さそり金融", 90000, date(2026, 7, 31), order_id="S1"),
        DemandItem("MOT2", 90000, date(2026, 7, 31), order_id="S2"),
    ]
    result = plan_bottleneck(demands, days, CAPS, stage_flows=flows, product_caps_by_mode=caps)

    # our の MIL 日次を機種別に取り出す
    our_mil = {}
    for c in result.stage_allocation:
        if c.stage_id == "MIL":
            our_mil.setdefault(c.product, {})[c.day] = c.quantity

    # FeliCa: さそり金融は our と同一形状(重複窓でMAE 0)、
    # MOT2 は同じ日だが台数を半分に(重複窓で台数ズレ → MAE > 0)
    sasori_co = dict(our_mil["さそり金融"])
    mot_co = {d: q / 2 for d, q in our_mil["MOT2"].items()}
    # Line-Inは空にする(この検査は完成日次形状のみ対象。同日Line-In+Completionの
    # carryover除外に引っかからないようにする)。
    buf = _felica_workbook([
        ("S1", "RC-SA02F/5", 90000, {}, sasori_co),
        ("S2", "RC-S127/HCB5", 90000, {}, mot_co),
    ])
    felica = parse_felica_plan(buf)

    aliases = {"RC-SA02F": "さそり金融", "RC-S127": "MOT2"}
    rep = compare_plans(result, felica, days, aliases=aliases)
    assert set(rep.daily_shape_by_product) == {"さそり金融", "MOT2"}
    assert rep.daily_shape_by_product["さそり金融"]["completion_mae"] == 0
    assert rep.daily_shape_by_product["MOT2"]["completion_mae"] > 0
    # aliases 省略時は機種別内訳を作らない(後方互換)
    assert compare_plans(result, felica, days).daily_shape_by_product == {}


def test_timing_by_product_reports_signed_bias():
    """機種別 予実タイミング差: ourが実績より遅い機種は completion_bias>0, 早い機種は<0。"""
    days = working_days_in_range(date(2026, 7, 1), date(2026, 8, 31))
    flows = [StageFlowConfig("ANT", -1), StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 1)]
    caps = {
        "16h": {"さそり金融": 90000, "MOT2": 90000},
        "22h": {"さそり金融": 120000, "MOT2": 120000},
    }
    demands = [
        DemandItem("さそり金融", 90000, date(2026, 7, 31), order_id="S1"),
        DemandItem("MOT2", 90000, date(2026, 7, 31), order_id="S2"),
    ]
    result = plan_bottleneck(demands, days, CAPS, stage_flows=flows, product_caps_by_mode=caps)
    our_comp = {lot.order_id: lot.completion_day for lot in result.mil_lots}

    ci_s1 = days.index(our_comp["S1"])
    ci_s2 = days.index(our_comp["S2"])
    # さそり金融(S1): FeliCa完成を our より2稼働日「前」に置く → our遅い → bias +2
    # MOT2(S2):     FeliCa完成を our より2稼働日「後」に置く → our早い → bias -2
    fel_s1 = days[ci_s1 - 2]
    fel_s2 = days[ci_s2 + 2]
    # Line-Inは空(完成日タイミングのみ検査。同日Line-In+Completionのcarryover除外回避)。
    buf = _felica_workbook([
        ("S1", "RC-SA02F/5", 90000, {}, {fel_s1: 90000}),
        ("S2", "RC-S127/HCB5", 90000, {}, {fel_s2: 90000}),
    ])
    felica = parse_felica_plan(buf)

    aliases = {"RC-SA02F": "さそり金融", "RC-S127": "MOT2"}
    rep = compare_plans(result, felica, days, aliases=aliases)
    assert rep.timing_by_product["さそり金融"]["completion_bias"] == 2.0
    assert rep.timing_by_product["MOT2"]["completion_bias"] == -2.0
    assert rep.timing_by_product["さそり金融"]["n"] == 1
    # aliases 省略時は機種別タイミングを作らない(後方互換)
    assert compare_plans(result, felica, days).timing_by_product == {}


def test_derive_stage_offsets_from_span_and_ratio():
    days = working_days_in_range(date(2026, 7, 1), date(2026, 7, 31))
    flows = [StageFlowConfig("ANT", -2), StageFlowConfig("TAL", -1),
             StageFlowConfig("HAL", 0), StageFlowConfig("MIL", 1)]  # 総スパン3, 上流2/下流1
    # 機種P: 投入7/1→完成7/6 = 3稼働日スパン(7/1,2,3,6)。
    # 機種Q(MOT2): 投入=完成が同日(スパン0)=先週までの計画台数(carryover)なので除外される。
    rows = [
        ("S1", "RC-SA02F/5", 90000, {date(2026, 7, 1): 90000}, {date(2026, 7, 6): 90000}),
        ("S2", "RC-SA02F/5", 90000, {date(2026, 7, 2): 90000}, {date(2026, 7, 7): 90000}),  # 3稼働日
        ("S3", "RC-S127/HCB5", 50000, {date(2026, 7, 1): 50000}, {date(2026, 7, 1): 50000}),  # 同日→除外
        ("S4", "RC-S127/HCB5", 50000, {date(2026, 7, 2): 50000}, {date(2026, 7, 2): 50000}),  # 同日→除外
    ]
    felica = parse_felica_plan(_felica_workbook(rows))
    derived = derive_stage_offsets(felica, days, flows)

    # さそり金融(RC-SA02F, スパン3) → ANT-2/TAL-1/MIL+1(現状比率どおり)
    assert derived["ANT"]["さそり金融"] == -2
    assert derived["MIL"]["さそり金融"] == 1
    # MOT2(同日Line-In+Completion=carryover)は除外され derived に現れない
    assert "MOT2" not in derived.get("ANT", {})
    assert "MOT2" not in derived.get("MIL", {})


def test_parse_felica_nonworking_days_reads_gray_weekday_cells():
    """行3の日付ヘッダーで gray125 塗り=非稼働日。週末は除外、平日の非稼働日だけ返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "202607_CTA1"
    gray = PatternFill(patternType="gray125")
    solid = PatternFill(patternType="solid", fgColor="FFFFFF")
    # col9以降に7/1..7/6を置く。7/4(土)7/5(日)=灰(週末), 7/2(木)=灰(祝日相当), 他=塗り
    header = [
        (date(2026, 7, 1), solid),
        (date(2026, 7, 2), gray),   # 平日の非稼働日 → 返る
        (date(2026, 7, 3), solid),
        (date(2026, 7, 4), gray),   # 土曜(週末) → 返らない
        (date(2026, 7, 5), gray),   # 日曜(週末) → 返らない
        (date(2026, 7, 6), solid),
    ]
    for i, (d, fill) in enumerate(header):
        cell = ws.cell(row=3, column=9 + i, value=d)
        cell.fill = fill
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    days = parse_felica_nonworking_days(buf)
    assert days == [date(2026, 7, 2)]  # 平日の灰のみ、週末は除外
