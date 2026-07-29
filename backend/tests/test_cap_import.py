import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from cap_import import (  # noqa: E402
    derive_product_caps,
    parse_machine_master,
    parse_product_aliases,
    parse_product_daily_caps,
    parse_recipe_codes,
)

# 現物の THM シートと同じ形: ■<n>h稼働 の下にヘッダー行(号機名)、その2行下から機種行、
# 最後に「設備キャパ」行。列は 1=呼称 2=モジュール 3=カード内製 4=外部委託 5..=号機。
MACHINES = ["ANT#2", "アンテナ検査機", "TAL#2", "HAL#7", "HAL#5", "MIL#3"]


def _cap_workbook(blocks):
    """blocks: {mode: {product: ([marks], 機種別キャパ)}} + 号機キャパ行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "THM"
    r = 1
    for mode, (rows, machine_caps) in blocks.items():
        ws.cell(row=r, column=1, value=f"■{mode}稼働")
        header = r + 1
        for i, name in enumerate(MACHINES):
            ws.cell(row=header, column=5 + i, value=name)
        ws.cell(row=header, column=5 + len(MACHINES), value="機種別\nキャパ")
        rr = header + 2
        for product, (marks, cap) in rows.items():
            ws.cell(row=rr, column=1, value=product)
            for i, mark in enumerate(marks):
                ws.cell(row=rr, column=5 + i, value=mark)
            ws.cell(row=rr, column=5 + len(MACHINES), value=cap)
            rr += 1
        ws.cell(row=rr, column=4, value="設備キャパ")
        for i, cap in enumerate(machine_caps):
            ws.cell(row=rr, column=5 + i, value=cap)
        r = rr + 3

    ms = wb.create_sheet("機種一覧")
    for c, label in enumerate(["呼称", "モデル名", "TALレシピ", "HALレシピ", "MILレシピ"], 1):
        ms.cell(row=1, column=c, value=label)
    for i, (product, model, recipe) in enumerate(
        [("さそり金融", "RC-SA02F/5", "A02F"),
         ("さそり金融", "RC-S100/HKT5", "A02F"),
         ("Lite-S(Mies)", "RC-S982F/5", "982F")], start=2
    ):
        ms.cell(row=i, column=1, value=product)
        ms.cell(row=i, column=2, value=model)
        for c in (3, 4, 5):
            ms.cell(row=i, column=c, value=recipe)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


BLOCKS = {
    "22h": (
        {   # ANT#2, 検査, TAL#2, HAL#7, HAL#5, MIL#3
            "さそり金融": (["○", "〇", "○", "○", "○", "○"], 60000),
            "Lite-S(Mies)": (["○", "〇", "○", "○", "×", "×"], 30000),
        },
        [144000, 140000, 80000, 30000, 30000, 42240],
    ),
    "16h": (
        {
            "さそり金融": (["○", "〇", "○", "○", "△"], 40000),
            "Lite-S(Mies)": (["○", "〇", "○", "×", "×"], 20000),
        },
        [108000, 100000, 60000, 20000, 20000, 30720],
    ),
}


def test_parse_machine_master_reads_stage_capacity_and_eligibility():
    master = parse_machine_master(_cap_workbook(BLOCKS))
    assert set(master) == {"22h", "16h"}

    by_id = {m.machine_id: m for m in master["22h"]}
    # 「アンテナ検査機」は工程が引けないので号機として採らない
    assert set(by_id) == {"ANT#2", "TAL#2", "HAL#7", "HAL#5", "MIL#3"}
    assert by_id["HAL#7"].stage_id == "HAL"
    assert by_id["HAL#7"].daily_capacity == 30000
    assert by_id["HAL#7"].eligibility == {"さそり金融": "○", "Lite-S(Mies)": "○"}
    # × は入らない
    assert by_id["HAL#5"].eligibility == {"さそり金融": "○"}
    assert by_id["HAL#5"].can_run("さそり金融")
    assert not by_id["HAL#5"].can_run("Lite-S(Mies)")


def test_conditional_mark_is_kept_and_can_be_excluded():
    master = parse_machine_master(_cap_workbook(BLOCKS))
    hal5 = next(m for m in master["16h"] if m.machine_id == "HAL#5")
    assert hal5.eligibility["さそり金融"] == "△"
    assert hal5.can_run("さそり金融")  # 既定では条件付きも使える
    assert not hal5.can_run("さそり金融", allow_conditional=False)


def test_parse_product_daily_caps_by_mode():
    caps = parse_product_daily_caps(_cap_workbook(BLOCKS))
    assert caps["22h"] == {"さそり金融": 60000, "Lite-S(Mies)": 30000}
    assert caps["16h"] == {"さそり金融": 40000, "Lite-S(Mies)": 20000}


def test_derive_product_caps_takes_the_narrowest_stage():
    master = parse_machine_master(_cap_workbook(BLOCKS))
    derived = derive_product_caps(master["22h"])
    # Lite-S は HAL#7 の1台(30,000)が最も細い。MIL#3は×なのでMILは0台=対象外。
    assert derived["Lite-S(Mies)"] == 30000
    # さそり金融は HAL 2台=60,000 が最小(TAL 80,000 / MIL 42,240 より…MILが42,240で最小)
    assert derived["さそり金融"] == 42240


def test_parse_product_aliases_registers_full_names_and_root_codes():
    """枝番付きのフル名だけだと台帳側の短い表記を最長前方一致で引けない。"""
    aliases = parse_product_aliases(_cap_workbook(BLOCKS))
    assert aliases["RC-SA02F/5"] == "さそり金融"
    assert aliases["RC-S100/HKT5"] == "さそり金融"
    # 枝番を落としたルートコードも登録される
    assert aliases["RC-S100"] == "さそり金融"
    assert aliases["RC-S982F"] == "Lite-S(Mies)"


def test_parse_recipe_codes_maps_stage_recipe_to_product():
    codes = parse_recipe_codes(_cap_workbook(BLOCKS))
    assert codes["A02F"] == "さそり金融"
    assert codes["982F"] == "Lite-S(Mies)"


def test_only_machines_filters_to_the_operating_set():
    """CAP表には稼働していない号機やモードで食い違うラベルの列が残っているので絞れる。"""
    master = parse_machine_master(_cap_workbook(BLOCKS), only_machines={"HAL#7", "MIL#3"})
    assert {m.machine_id for m in master["22h"]} == {"HAL#7", "MIL#3"}
    # 絞っても能力・可否は変わらない
    hal7 = next(m for m in master["22h"] if m.machine_id == "HAL#7")
    assert hal7.daily_capacity == 30000
    assert hal7.eligibility == {"さそり金融": "○", "Lite-S(Mies)": "○"}


def test_active_machines_default_covers_the_real_line():
    """既定の稼働号機一覧(TAL#2/#3, HAL#5-#9, MIL#3/#5/#6/#7)。"""
    from cap_import import ACTIVE_MACHINES

    assert ACTIVE_MACHINES == frozenset({
        "TAL#2", "TAL#3",
        "HAL#5", "HAL#6", "HAL#7", "HAL#8", "HAL#9",
        "MIL#3", "MIL#5", "MIL#6", "MIL#7",
    })
    # ANT号機は持たない(工程展開のリードタイムだけで表現する)
    assert not any(m.startswith("ANT") for m in ACTIVE_MACHINES)
