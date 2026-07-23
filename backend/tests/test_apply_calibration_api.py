import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from fastapi.testclient import TestClient  # noqa: E402

import main  # noqa: E402
from config_loader import CONFIG_DIR  # noqa: E402

client = TestClient(main.app)
CFG = CONFIG_DIR / "bottleneck_planning.json"


def test_apply_calibration_writes_offsets_then_restores():
    """推奨較正値を config に反映するAPI(実configをバックアップ→検証→復元)。"""
    original = CFG.read_bytes()
    try:
        resp = client.post(
            "/api/bottleneck/apply-calibration",
            json={"offsets": {"ANT": -3, "TAL": -2, "HAL": 0, "MIL": 1}, "a_shift_fraction": 0.6},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["applied"] is True
        assert body["offsets"] == {"ANT": -3, "TAL": -2, "HAL": 0, "MIL": 1}
        assert body["a_shift_fraction"] == 0.6

        data = json.loads(CFG.read_text(encoding="utf-8"))
        flows = {f["stageId"]: f["leadOffsetDays"] for f in data["stageFlows"]}
        assert flows == {"ANT": -3, "TAL": -2, "HAL": 0, "MIL": 1}
        assert data["aShiftFraction"] == 0.6
        # 機種別キャパ等が保持されている(反映が破壊的でない)
        assert "productDailyCapsByMode" in data
    finally:
        CFG.write_bytes(original)


def test_apply_calibration_rejects_bad_stage_key():
    resp = client.post(
        "/api/bottleneck/apply-calibration",
        json={"offsets": {"XXX": 0}, "a_shift_fraction": 0.5},
    )
    assert resp.status_code == 422


def test_apply_calibration_rejects_out_of_range_fraction():
    resp = client.post(
        "/api/bottleneck/apply-calibration",
        json={"offsets": {"ANT": -1, "TAL": -2, "HAL": 0, "MIL": 2}, "a_shift_fraction": 1.5},
    )
    assert resp.status_code == 422


def test_apply_calibration_rejects_extreme_offset():
    resp = client.post(
        "/api/bottleneck/apply-calibration",
        json={"offsets": {"ANT": -99, "TAL": -2, "HAL": 0, "MIL": 2}, "a_shift_fraction": 0.5},
    )
    assert resp.status_code == 422


def _felica_calendar_bytes():
    """gray125(非稼働日)塗りの日付ヘッダーを持つ最小FeliCaブックのbytesを返す。"""
    import io
    from datetime import date

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "202608_CTA1"
    gray = PatternFill(patternType="gray125")
    solid = PatternFill(patternType="solid", fgColor="FFFFFF")
    header = [(date(2026, 8, 10), solid), (date(2026, 8, 11), gray), (date(2026, 8, 12), solid)]
    for i, (d, fill) in enumerate(header):
        c = ws.cell(row=3, column=9 + i, value=d)
        c.fill = fill
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_apply_calendar_writes_nonworking_days_then_restores():
    original = CFG.read_bytes()
    try:
        resp = client.post(
            "/api/bottleneck/apply-calendar",
            files={"felica_file": ("f.xlsx", _felica_calendar_bytes(),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["applied"] is True
        assert body["non_working_days"] == ["2026-08-11"]  # 平日の灰のみ
        data = json.loads(CFG.read_text(encoding="utf-8"))
        assert data["nonWorkingDays"] == ["2026-08-11"]
        assert "productDailyCapsByMode" in data  # 既存フィールド保持
    finally:
        CFG.write_bytes(original)


def test_apply_calendar_rejects_bad_file():
    resp = client.post(
        "/api/bottleneck/apply-calendar",
        files={"felica_file": ("bad.xlsx", b"not-a-workbook", "application/octet-stream")},
    )
    assert resp.status_code == 400
