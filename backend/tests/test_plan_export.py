import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from config_loader import load_changeover_config, load_equipment_config, load_orders_data  # noqa: E402
from plan_export import export_plan_workbook  # noqa: E402
from scheduler import Scheduler  # noqa: E402


def _run_plan():
    equipment = load_equipment_config()
    changeover = load_changeover_config()
    orders = load_orders_data()
    result = Scheduler(equipment, changeover, orders).run()
    return result, equipment


def _reload(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_workbook_has_expected_sheets_and_no_utilization():
    result, equipment = _run_plan()
    wb = _reload(export_plan_workbook(result, equipment))
    assert wb.sheetnames == ["サマリー", "スケジュール明細", "日付×シフト台数", "警告"]
    # 稼働率シートは含めない
    assert not any("稼働" in name for name in wb.sheetnames)


def test_schedule_sheet_row_count_matches_schedule():
    result, equipment = _run_plan()
    wb = _reload(export_plan_workbook(result, equipment))
    ws = wb["スケジュール明細"]
    # ヘッダー1行 + 作業数
    assert ws.max_row == len(result.schedule) + 1
    headers = [c.value for c in ws[1]]
    assert headers[0] == "受注ID"
    assert "段取替え(分)" in headers


def test_matrix_sheet_counts_split_lots_per_stage():
    result, equipment = _run_plan()
    wb = _reload(export_plan_workbook(result, equipment))
    ws = wb["日付×シフト台数"]
    # 2段ヘッダー(日付/シフト名)の下に受注行が並ぶ
    order_labels = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
    order_labels = [v for v in order_labels if v and v.startswith("ORD-")]
    assert len(order_labels) == len({op.order_id for op in result.schedule})

    # いずれかのセルに "1/4/4" のような工程別台数(分割4リール)が現れることを確認
    found_four = False
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if isinstance(cell.value, str) and "/" in cell.value:
                parts = cell.value.split("/")
                if len(parts) == 3 and (parts[1] == "4" or parts[2] == "4"):
                    found_four = True
    assert found_four, "分割ロット(工程B/Cが4台)のセルが見つかりません"


def test_warnings_sheet_lists_warnings():
    result, equipment = _run_plan()
    wb = _reload(export_plan_workbook(result, equipment))
    ws = wb["警告"]
    assert [c.value for c in ws[1]] == ["対象", "メッセージ"]
    # サンプルデータでは安全在庫割れ警告が1件以上出る
    body = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert len(body) >= 1
