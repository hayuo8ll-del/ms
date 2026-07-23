import io
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from config_loader import (  # noqa: E402
    load_bottleneck_planning,
    save_bottleneck_calibration,
    save_nonworking_days,
)
from thm_ledger_import import (  # noqa: E402
    parse_actuals,
    parse_daily_actuals,
    parse_equipment_stops,
    parse_ta1_hal_actuals,
    parse_thm_ledger,
    parse_thm_shortterm_actuals,
    resolve_product,
)


def test_resolve_product_uses_longest_prefix_match():
    # RC-コードの最長一致で呼称を解決する
    assert resolve_product("RC-SA02F/5  J") == "さそり金融"
    assert resolve_product("RC-S100/HNA5HK") == "さそり金融"
    assert resolve_product("RC-SA05A/5") == "さそり交通"
    assert resolve_product("RC-S105/JE16 J INLAY") == "SuicaⅢ"
    assert resolve_product("RC-S982F/5 J") == "Lite-S(Mies)"
    assert resolve_product("RC-SA10A J") == "部分リライト"  # スラッシュ無し・末尾サフィックス
    assert resolve_product("RC-S140     WW") == "SD3"
    assert resolve_product("まったく不明な品目") is None


def _ledger_workbook(rows):
    """台帳形式(1行目空・2行目ヘッダー・3行目以降データ)のブックを作る。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "台帳"
    headers = ["№", "ライン", "完成品名", "x", "完成品コード", "製番", "ICチップ", "y", "着手予定日", "完成予定日", "完成予定数", "出荷日"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    for i, row in enumerate(rows, start=3):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_parse_ledger_uses_seiban_as_order_id():
    rows = [
        ["THM1", "CTA1", "RC-SA02F/5  J", None, "98x", "6F5Z9FRY", "東芝さそり（6K8pF）", None, None, date(2026, 7, 15), 12000],
        ["THM2", "CTA2", "RC-SA05A/5", None, "98y", "6RAZ9EEO", "東芝さそり（6K7pF）", None, None, date(2026, 7, 20), 8000],
        ["THM3", "CRC1", "謎の品目", None, "98z", "s3", "?", None, None, date(2026, 7, 22), 5000],
    ]
    demands, unmapped = parse_thm_ledger(_ledger_workbook(rows))

    assert {d.product for d in demands} == {"さそり金融", "さそり交通"}
    # 出荷ロットの識別子は製番列(現場MIL表と同じキー)
    assert {d.order_id for d in demands} == {"6F5Z9FRY", "6RAZ9EEO"}
    assert demands[0].quantity == 12000
    assert demands[0].due_date == date(2026, 7, 15)
    # マップできなかった行は unmapped に載る
    assert len(unmapped) == 1 and unmapped[0].order_id == "s3"


def test_parse_ledger_falls_back_to_no_when_seiban_empty():
    rows = [
        ["THM9", "CTA1", "RC-SA02F/5", None, None, "-", None, None, None, date(2026, 7, 15), 100],
    ]
    demands, _ = parse_thm_ledger(_ledger_workbook(rows))
    assert demands[0].order_id == "THM9"


def test_parse_ledger_filters_future_due_and_lines():
    rows = [
        ["THM1", "CTA1", "RC-SA02F/5", None, None, None, None, None, None, date(2026, 6, 30), 100],  # 過去納期
        ["THM2", "CTA1", "RC-SA02F/5", None, None, None, None, None, None, date(2026, 7, 25), 200],  # 未来・CTA1
        ["THM3", "CRC1", "RC-SA02F/5", None, None, None, None, None, None, date(2026, 7, 25), 300],  # 未来・CRC1
    ]
    demands, _ = parse_thm_ledger(
        _ledger_workbook(rows),
        only_due_on_or_after=date(2026, 7, 21),
        lines={"CTA1", "CTA2"},
    )
    assert [d.order_id for d in demands] == ["THM2"]


def test_due_date_is_ship_date_minus_buffer():
    # 出荷日=真の納期。完成目標=出荷日−2日(暦日, 既定)。完成予定日(7/15)は使わない。
    rows = [
        ["THM1", "CTA1", "RC-SA02F/5", None, None, "S1", None, None, None, date(2026, 7, 15), 12000, date(2026, 8, 7)],
    ]
    demands, _ = parse_thm_ledger(_ledger_workbook(rows))
    d = demands[0]
    assert d.ship_date == date(2026, 8, 7)
    assert d.due_date == date(2026, 8, 5)  # 8/7 − 2日


def test_buffer_is_configurable_and_falls_back_to_completion_when_no_ship():
    rows = [
        ["THM1", "CTA1", "RC-SA02F/5", None, None, "S1", None, None, None, date(2026, 7, 15), 100, date(2026, 8, 10)],
        ["THM2", "CTA1", "RC-SA05A/5", None, None, "S2", None, None, None, date(2026, 7, 20), 200, None],  # 出荷日なし
    ]
    demands, _ = parse_thm_ledger(_ledger_workbook(rows), shipment_buffer_days=3)
    by = {d.order_id: d for d in demands}
    assert by["S1"].due_date == date(2026, 8, 7)   # 8/10 − 3日
    assert by["S2"].ship_date is None
    assert by["S2"].due_date == date(2026, 7, 20)  # 出荷日なし → 完成予定日にフォールバック(バッファ無し)


def _with_actuals_sheet(buf, rows):
    """既存ブックに「実績」シート(製番/実績数)を追加して返す。"""
    from openpyxl import load_workbook

    wb = load_workbook(buf)
    ws = wb.create_sheet("実績")
    ws.append(["製番", "実績数"])
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def test_parse_actuals_reads_seiban_quantities_and_sums_duplicates():
    base = _ledger_workbook([])
    buf = _with_actuals_sheet(base, [["6F5Z9FRY", 5000], ["6F5Z9FRY", 2000], ["6RAZ9EEO", 800], ["", 999]])
    actuals = parse_actuals(buf)
    assert actuals == {"6F5Z9FRY": 7000.0, "6RAZ9EEO": 800.0}


def test_parse_actuals_returns_empty_when_sheet_missing():
    assert parse_actuals(_ledger_workbook([])) == {}


def test_parse_daily_actuals_sums_per_day_across_stages():
    base = _ledger_workbook([])
    from openpyxl import load_workbook

    wb = load_workbook(base)
    ws = wb.create_sheet("日次実績")
    ws.append(["日付", "工程", "実績数"])
    ws.append([date(2026, 7, 21), "HAL", 50000])
    ws.append([date(2026, 7, 21), "HAL", 30000])  # 同日は合算
    ws.append([date(2026, 7, 22), "HAL", 90000])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    daily = parse_daily_actuals(out)
    assert daily == {date(2026, 7, 21): 80000.0, date(2026, 7, 22): 90000.0}


def test_parse_daily_actuals_empty_when_sheet_missing():
    assert parse_daily_actuals(_ledger_workbook([])) == {}


def test_parse_equipment_stops_reads_enabled_rows_only():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "設備停止"
    headers = ["停止ID", "有効", "停止区分", "工程", "設備", "開始日", "開始勤務", "終了日", "終了勤務",
               "停止率_%", "停止時間_h", "補正後Cap_台", "停止Cap控除方法"]
    ws.append(headers)
    ws.append(["S0001", "Y", "オーバーホール", "HAL", "HAL#9", date(2026, 7, 22), "A勤", date(2026, 7, 24), "B勤",
               100, 0, None, "全停止"])
    ws.append(["S0002", "N", "試作", "HAL", "HAL#8", date(2026, 7, 20), "A勤", date(2026, 7, 20), "B勤",
               100, 0, None, "全停止"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stops = parse_equipment_stops(buf)
    assert len(stops) == 1
    s = stops[0]
    assert (s.stop_id, s.stage_id, s.machine_id, s.method) == ("S0001", "HAL", "HAL#9", "全停止")
    assert (s.start_day, s.end_day, s.start_shift, s.end_shift) == (date(2026, 7, 22), date(2026, 7, 24), "A勤", "B勤")


def test_parse_equipment_stops_empty_when_sheet_missing():
    assert parse_equipment_stops(_ledger_workbook([])) == []


def test_load_bottleneck_planning_reads_config_file():
    cfg = load_bottleneck_planning()
    # config/bottleneck_planning.json の内容(≒組み込み既定値)が読める
    assert cfg.line_daily_capacities["16h"] == 90000
    assert cfg.bottleneck_stage == "HAL"
    assert cfg.stage_order == ["ANT", "TAL", "HAL", "MIL"]
    assert cfg.product_daily_caps_by_mode["16h"]["Lite-S(Mies)"] == 30720
    assert cfg.product_aliases["RC-SA02F"] == "さそり金融"
    assert cfg.machine_counts["HAL"] == 5


def test_save_bottleneck_calibration_updates_offsets_and_preserves_rest(tmp_path):
    import json

    cfg_path = tmp_path / "bottleneck_planning.json"
    cfg_path.write_text(
        json.dumps(
            {
                "lineDailyCapacities": {"16h": 90000, "22h": 120000},
                "bottleneckStage": "HAL",
                "stageFlows": [
                    {"stageId": "ANT", "leadOffsetDays": -2},
                    {"stageId": "TAL", "leadOffsetDays": -1, "inputUnit": 40000},
                    {"stageId": "HAL", "leadOffsetDays": 0, "inputUnit": 10000},
                    {"stageId": "MIL", "leadOffsetDays": 1, "inputUnit": 1920},
                ],
                "aShiftFraction": 0.5,
                "productDailyCapsByMode": {"16h": {"さそり金融": 80000}},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    saved = save_bottleneck_calibration({"ANT": -1, "TAL": -2, "HAL": 0, "MIL": 2}, 0.4, path=cfg_path)
    assert saved["offsets"] == {"ANT": -1, "TAL": -2, "HAL": 0, "MIL": 2}
    assert saved["a_shift_fraction"] == 0.4

    data = json.loads(cfg_path.read_text(encoding="utf-8"))
    flows = {f["stageId"]: f for f in data["stageFlows"]}
    # オフセット/A勤割合が更新される
    assert flows["ANT"]["leadOffsetDays"] == -1
    assert flows["MIL"]["leadOffsetDays"] == 2
    assert data["aShiftFraction"] == 0.4
    # inputUnit・機種別キャパ等は保持される
    assert flows["TAL"]["inputUnit"] == 40000
    assert flows["MIL"]["inputUnit"] == 1920
    assert data["productDailyCapsByMode"]["16h"]["さそり金融"] == 80000


def test_save_and_load_nonworking_days(tmp_path):
    """FeliCa由来の非稼働日を書き込み→再読込できる(他フィールド保持・昇順)。"""
    import json
    from datetime import date

    cfg_path = tmp_path / "bottleneck_planning.json"
    cfg_path.write_text(
        json.dumps({"aShiftFraction": 0.4, "productAliases": {"RC-S100": "さそり金融"}}, ensure_ascii=False),
        encoding="utf-8",
    )
    iso = save_nonworking_days([date(2026, 8, 14), date(2026, 7, 20), date(2026, 8, 11)], path=cfg_path)
    assert iso == ["2026-07-20", "2026-08-11", "2026-08-14"]
    data = json.loads(cfg_path.read_text(encoding="utf-8"))
    assert data["nonWorkingDays"] == ["2026-07-20", "2026-08-11", "2026-08-14"]
    assert data["aShiftFraction"] == 0.4
    assert data["productAliases"]["RC-S100"] == "さそり金融"


def test_parse_thm_shortterm_actuals_sums_red_font_per_stage():
    """THM短期投入予定表: Line-In(TAL)/Completion(MIL)行の赤字=実績だけを製番別に合算。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    red = Font(color="FFFF0000")
    black = Font(color="FF000000")
    wb = Workbook()
    ws = wb.active
    ws.title = "TA1"
    # 製番S1: Line-In行(r5) TAL実績=1000(赤)+予定500(黒), Completion行(r6) MIL実績=800+200(赤)
    ws.cell(row=5, column=3, value="S1")
    ws.cell(row=5, column=6, value="Line-In")
    ws.cell(row=5, column=7, value=1000).font = red
    ws.cell(row=5, column=8, value=500).font = black  # 予定→無視
    ws.cell(row=6, column=7, value=800).font = red
    ws.cell(row=6, column=9, value=200).font = red
    # 製番S2: 実績なし(全部黒) → 出力に含まれない
    ws.cell(row=7, column=3, value="S2")
    ws.cell(row=7, column=6, value="Line-In")
    ws.cell(row=7, column=7, value=300).font = black
    ws.cell(row=8, column=7, value=300).font = black
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    out = parse_thm_shortterm_actuals(buf)
    assert out == {"S1": {"TAL": 1000.0, "MIL": 1000.0}}  # MIL=800+200


def test_parse_ta1_hal_actuals_reads_red_hal_with_month_rollover():
    """TA1_投入計画: HAL行の赤字=実績を日別に。月が戻ったら年を+1(年跨ぎ)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    red = Font(color="FFFF0000")
    black = Font(color="FF000000")
    wb = Workbook()
    ws = wb.active
    ws.title = "生産計画"
    # 行1: 月ラベル(持ち越し), 行2: 日, HAL行の赤字を拾う
    ws.cell(row=1, column=3, value="12月")
    ws.cell(row=1, column=5, value="1月")   # 12→1 で年+1
    ws.cell(row=2, column=3, value=30)
    ws.cell(row=2, column=4, value=31)
    ws.cell(row=2, column=5, value=1)
    ws.cell(row=2, column=6, value=2)
    ws.cell(row=6, column=1, value="HAL")
    ws.cell(row=6, column=3, value=100).font = red
    ws.cell(row=6, column=4, value=200).font = red
    ws.cell(row=6, column=5, value=300).font = red
    ws.cell(row=6, column=6, value=400).font = black  # 予定→無視
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    hal = parse_ta1_hal_actuals(buf, year=2026)
    assert hal[date(2026, 12, 30)] == 100.0
    assert hal[date(2026, 12, 31)] == 200.0
    assert hal[date(2027, 1, 1)] == 300.0  # 年跨ぎ
    assert date(2027, 1, 2) not in hal      # 黒字は実績でない
