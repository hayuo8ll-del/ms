import io
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from bottleneck_planner import (  # noqa: E402
    DemandItem,
    StageFlowConfig,
    plan_bottleneck,
    working_days_in_range,
)
from bottleneck_export import export_bottleneck_workbook  # noqa: E402

CAPS = {"16h": 90000, "22h": 120000}
STAGE_ORDER = ["ANT", "TAL", "HAL", "MIL"]


def _make_result():
    days = working_days_in_range(date(2026, 7, 1), date(2026, 7, 31))
    demands = [
        DemandItem("さそり金融", 90000, date(2026, 7, 10), order_id="L1"),
        DemandItem("SuicaⅢ", 90000, date(2026, 7, 1), order_id="L2"),  # 納期7/1・MIL完成7/2→超過
    ]
    flows = [StageFlowConfig(s, off) for s, off in [("ANT", -2), ("TAL", -1), ("HAL", 0), ("MIL", 1)]]
    result = plan_bottleneck(demands, days, CAPS, stage_flows=flows)
    return result, demands


def _reload(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_workbook_has_the_two_shopfloor_sheets():
    result, demands = _make_result()
    wb = _reload(export_bottleneck_workbook(result, demands, STAGE_ORDER))
    assert wb.sheetnames == ["サマリー", "生産計画(機種×日)", "製番別MIL", "段取り", "警告"]


def test_changeover_sheet_lists_campaigns_with_switch_flag():
    result, demands = _make_result()
    wb = _reload(export_bottleneck_workbook(result, demands, STAGE_ORDER))
    ws = wb["段取り"]
    assert [ws.cell(row=1, column=c).value for c in range(1, 8)] == [
        "工程", "機種", "開始日", "終了日", "日数", "数量", "区分"
    ]
    kinds = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    # 立上げ(工程初日)と切替(別機種からの段取り)の両方が現れる
    assert "立上げ" in kinds
    assert "切替(段取り)" in kinds
    # 行数 = キャンペーン数
    assert ws.max_row - 1 == len(result.campaigns)


def test_stage_matrix_lists_each_product_with_all_stages():
    result, demands = _make_result()
    wb = _reload(export_bottleneck_workbook(result, demands, STAGE_ORDER))
    ws = wb["生産計画(機種×日)"]
    # 工程列(col2)に ANT/TAL/HAL/MIL が機種数ぶん現れる
    stages = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert stages.count("HAL") == 2  # 2機種
    assert set(STAGE_ORDER).issubset(set(stages))


def test_mil_sheet_has_one_row_per_lot_and_flags_late():
    result, demands = _make_result()
    wb = _reload(export_bottleneck_workbook(result, demands, STAGE_ORDER))
    ws = wb["製番別MIL"]
    order_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert set(order_ids) == {"L1", "L2"}
    assert [ws.cell(row=1, column=c).value for c in (5, 6, 7)] == ["出荷日(納期)", "完成目標", "判定"]
    judges = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert any(j and j.startswith("×") for j in judges)  # 完成目標を超過するロット
