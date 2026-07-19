import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from config_loader import load_changeover_config, load_equipment_config, load_orders_data  # noqa: E402
from excel_import import (  # noqa: E402
    ImportValidationError,
    WorkbookReadError,
    export_workbook,
    parse_workbook,
)
from scheduler import Scheduler  # noqa: E402


def _wb_to_stream(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _minimal_workbook(**overrides) -> Workbook:
    """全必須シートを備えた最小構成のワークブックを返す。overridesでシート内容を差し替え可能。"""
    sheets = {
        "Stages": (
            ["stageId", "stageName", "order", "uninterruptible", "batchRounding"],
            [["S1", "工程1", 1, False, ""]],
        ),
        "Machines": (
            ["stageId", "machineId", "name", "capacityPerHour"],
            [["S1", "M1", "1号機", 100]],
        ),
        "LotSplitting": (["stageAfter", "splitInto"], [["", 1]]),
        "ShiftModes": (
            ["shiftMode", "shiftName", "start", "end"],
            [["8h", "A勤", "08:00", "16:00"]],
        ),
        "Settings": (
            ["key", "value"],
            [["defaultShiftMode", "8h"], ["planStart", "2026-07-20T08:00:00"]],
        ),
        "Changeover": (["stageId", "fromProduct", "toProduct", "minutes"], []),
        "AShiftOnlyTransitions": (["stageId", "fromProduct", "toProduct"], []),
        "Orders": (
            ["orderId", "product", "quantity", "dueDate"],
            [["O1", "X", 50, "2026-07-25"]],
        ),
        "Inventory": (["product", "currentStock", "safetyStock"], [["X", 100, 50]]),
        "RawMaterials": (["materialId", "product", "onHand"], [["MAT-X", "X", 1000]]),
        "RawMaterialIncoming": (["product", "date", "quantity"], []),
    }
    sheets.update(overrides)

    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    return wb


def test_minimal_workbook_parses_successfully_and_schedules():
    equipment, changeover, orders_doc, summary = parse_workbook(_wb_to_stream(_minimal_workbook()))

    assert summary.stages == 1
    assert summary.machines == 1
    assert summary.orders == 1
    assert equipment["stages"][0]["stageId"] == "S1"
    assert orders_doc["orders"][0]["orderId"] == "O1"

    # 取り込んだ内容がそのままスケジューラで実行できることを確認する
    import config_loader

    eq = config_loader.EquipmentConfig(
        stages=[
            config_loader.StageConfig(
                stage_id=s["stageId"],
                name=s["stageName"],
                order=s["order"],
                machines=[
                    config_loader.MachineConfig(m["machineId"], m["name"], m["capacityPerHour"])
                    for m in s["machines"]
                ],
                uninterruptible=s.get("uninterruptible", False),
                batch_rounding=s.get("batchRounding"),
            )
            for s in equipment["stages"]
        ],
        lot_split_after=equipment["lotSplitting"]["stageAfter"],
        lot_split_into=equipment["lotSplitting"]["splitInto"],
        shift_modes=equipment["shiftModes"],
        default_shift_mode=equipment["defaultShiftMode"],
    )
    co = config_loader.ChangeoverConfig(matrix=changeover, a_shift_only_transitions=changeover.get("aShiftOnlyTransitions", {}))
    from datetime import date as _date

    orders_data = config_loader.OrdersData(
        orders=[
            config_loader.Order(o["orderId"], o["product"], o["quantity"], _date.fromisoformat(o["dueDate"]))
            for o in orders_doc["orders"]
        ],
        inventory={
            p: config_loader.Inventory(v["currentStock"], v["safetyStock"]) for p, v in orders_doc["inventory"].items()
        },
        raw_materials={
            p: config_loader.RawMaterial(v["materialId"], v["onHand"], []) for p, v in orders_doc["rawMaterials"].items()
        },
        plan_start=__import__("datetime").datetime.fromisoformat(orders_doc["planStart"]),
    )
    result = Scheduler(eq, co, orders_data).run()
    assert len(result.schedule) == 1


def test_round_trip_export_then_import_matches_current_config():
    wb = export_workbook()
    equipment, changeover, orders_doc, summary = parse_workbook(_wb_to_stream(wb))

    expected_equipment = load_equipment_config()
    expected_orders = load_orders_data()

    assert summary.stages == len(expected_equipment.stages)
    assert summary.orders == len(expected_orders.orders)
    assert {s["stageId"] for s in equipment["stages"]} == {s.stage_id for s in expected_equipment.stages}
    assert {o["orderId"] for o in orders_doc["orders"]} == {o.order_id for o in expected_orders.orders}


def test_missing_required_sheet_is_reported():
    sheets = {
        "Stages": (["stageId", "stageName", "order", "uninterruptible", "batchRounding"], [["S1", "工程1", 1, False, ""]]),
        "Machines": (["stageId", "machineId", "name", "capacityPerHour"], [["S1", "M1", "1号機", 100]]),
        "LotSplitting": (["stageAfter", "splitInto"], [["", 1]]),
        "ShiftModes": (["shiftMode", "shiftName", "start", "end"], [["8h", "A勤", "08:00", "16:00"]]),
        "Settings": (["key", "value"], [["defaultShiftMode", "8h"], ["planStart", "2026-07-20T08:00:00"]]),
        "Changeover": (["stageId", "fromProduct", "toProduct", "minutes"], []),
        "Orders": (["orderId", "product", "quantity", "dueDate"], [["O1", "X", 50, "2026-07-25"]]),
        "Inventory": (["product", "currentStock", "safetyStock"], [["X", 100, 50]]),
        # RawMaterials シートを意図的に省略
    }
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    try:
        parse_workbook(_wb_to_stream(wb))
        assert False, "ImportValidationError が発生するはず"
    except ImportValidationError as e:
        assert any(i.sheet == "RawMaterials" for i in e.issues)


def test_invalid_numeric_field_is_reported_with_row_number():
    wb = _minimal_workbook(
        Machines=(["stageId", "machineId", "name", "capacityPerHour"], [["S1", "M1", "1号機", "たくさん"]])
    )
    try:
        parse_workbook(_wb_to_stream(wb))
        assert False, "ImportValidationError が発生するはず"
    except ImportValidationError as e:
        issue = next(i for i in e.issues if i.sheet == "Machines")
        assert issue.row == 2
        assert "capacityPerHour" in issue.message


def test_unknown_stage_reference_in_machines_is_reported():
    wb = _minimal_workbook(
        Machines=(["stageId", "machineId", "name", "capacityPerHour"], [["S999", "M1", "1号機", 100]])
    )
    try:
        parse_workbook(_wb_to_stream(wb))
        assert False, "ImportValidationError が発生するはず"
    except ImportValidationError as e:
        assert any("S999" in i.message for i in e.issues)


def test_duplicate_order_id_is_reported():
    wb = _minimal_workbook(
        Orders=(
            ["orderId", "product", "quantity", "dueDate"],
            [["O1", "X", 50, "2026-07-25"], ["O1", "X", 30, "2026-07-26"]],
        )
    )
    try:
        parse_workbook(_wb_to_stream(wb))
        assert False, "ImportValidationError が発生するはず"
    except ImportValidationError as e:
        assert any("重複" in i.message for i in e.issues)


def test_stage_with_no_machines_is_reported():
    wb = _minimal_workbook(Machines=(["stageId", "machineId", "name", "capacityPerHour"], []))
    try:
        parse_workbook(_wb_to_stream(wb))
        assert False, "ImportValidationError が発生するはず"
    except ImportValidationError as e:
        assert any("号機が1台も定義されていません" in i.message for i in e.issues)


def test_corrupted_file_raises_workbook_read_error():
    try:
        parse_workbook(io.BytesIO(b"not an excel file"))
        assert False, "WorkbookReadError が発生するはず"
    except WorkbookReadError:
        pass
