"""立案した生産計画(PlanResult)を1つのExcelワークブックに出力する。

Web画面の「計画をExcel出力」で使う。以下のシートを含む(稼働率は含めない):

  サマリー / スケジュール明細 / 日付×シフト台数 / 警告

日付×シフト台数シートは、画面の日付×シフト台数マトリクス(app.jsのrenderShiftMatrix)と
同じ集計(行=受注、列=日付×シフト、セル=工程A/B/C別の稼働号機数)をサーバー側で行う。
"""
from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import EquipmentConfig, PlanResult, ScheduledOp
from shift_calendar import ShiftCalendar

# 工程ごとのセル塗り色(画面の工程A/B/C配色に対応)。
_STAGE_FILLS = {
    "STAGE1": "DCE7FB",  # 工程A(青系)
    "STAGE2": "D6F0E6",  # 工程B(緑系)
    "STAGE3": "F3E7C9",  # 工程C(琥珀系)
}
_MATRIX_STAGES = ["STAGE1", "STAGE2", "STAGE3"]

_HEADER_FILL = PatternFill("solid", fgColor="2D3139")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_DT_FORMAT = "yyyy-mm-dd hh:mm"
_THIN = Side(style="thin", color="D0D4DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_active_windows(schedule: list[ScheduledOp], equipment: EquipmentConfig):
    """opがいずれか重なるシフトウィンドウ(start, end, name)のみを時系列で返す。"""
    starts = [op.start for op in schedule]
    ends = [op.end for op in schedule]
    min_dt, max_dt = min(starts), max(ends)
    horizon = (max_dt.date() - min_dt.date()).days + 3
    windows = ShiftCalendar._build_windows(equipment.active_shift_defs(), min_dt.date(), horizon)
    return [w for w in windows if any(op.start < w[1] and op.end > w[0] for op in schedule)]


def _add_summary_sheet(wb: Workbook, result: PlanResult) -> None:
    ws = wb.create_sheet("サマリー")
    order_count = len({op.order_id for op in result.schedule})
    rows = [
        ("計画開始日", result.plan_start.strftime("%Y-%m-%d %H:%M")),
        ("対象オーダー数", order_count),
        ("総作業数", len(result.schedule)),
        ("警告件数", len(result.warnings)),
    ]
    ws.append(["項目", "値"])
    _style_header_row(ws, 1, 2)
    for label, value in rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = _BOLD
    _set_widths(ws, [20, 26])
    ws.freeze_panes = "A2"


def _add_schedule_sheet(wb: Workbook, result: PlanResult) -> None:
    ws = wb.create_sheet("スケジュール明細")
    headers = ["受注ID", "ロットID", "工程", "号機", "製品", "数量", "開始", "終了", "段取替え(分)", "備考"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    # 開始時刻→工程順で並べ、読みやすくする
    ops = sorted(result.schedule, key=lambda o: (o.start, o.stage_id, o.machine_id))
    for op in ops:
        ws.append(
            [
                op.order_id,
                op.lot_id,
                op.stage_id,
                op.machine_id,
                op.product,
                op.quantity,
                op.start,
                op.end,
                op.changeover_minutes,
                op.note,
            ]
        )
        r = ws.max_row
        ws.cell(row=r, column=7).number_format = _DT_FORMAT
        ws.cell(row=r, column=8).number_format = _DT_FORMAT

    _set_widths(ws, [12, 14, 9, 8, 9, 9, 17, 17, 12, 34])
    ws.freeze_panes = "A2"


def _add_matrix_sheet(wb: Workbook, result: PlanResult, equipment: EquipmentConfig) -> None:
    ws = wb.create_sheet("日付×シフト台数")
    schedule = result.schedule
    if not schedule:
        ws.append(["スケジュールがありません。"])
        return

    windows = _build_active_windows(schedule, equipment)

    # (order, window_index, stage) -> 稼働号機の集合
    orders = sorted({op.order_id for op in schedule})
    product_of = {op.order_id: op.product for op in schedule}
    counts: dict[tuple[str, int, str], set[str]] = {}
    for op in schedule:
        for wi, (w_start, w_end, _name) in enumerate(windows):
            if op.start < w_end and op.end > w_start:
                counts.setdefault((op.order_id, wi, op.stage_id), set()).add(op.machine_id)

    # 1行目=日付(シフト列をまたいで結合), 2行目=シフト名。先頭列は「受注」。
    ws.cell(row=1, column=1, value="受注")
    ws.cell(row=2, column=1, value="")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    date_groups: list[tuple[str, int]] = []
    for w_start, _w_end, _name in windows:
        key = f"{w_start.month}/{w_start.day}"
        if date_groups and date_groups[-1][0] == key:
            date_groups[-1] = (key, date_groups[-1][1] + 1)
        else:
            date_groups.append((key, 1))
    # 日付ヘッダー(結合)
    c = 2
    for label, span in date_groups:
        ws.cell(row=1, column=c, value=label)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + span - 1)
        c += span
    # シフト名ヘッダー
    for wi, (_w_start, _w_end, name) in enumerate(windows):
        ws.cell(row=2, column=2 + wi, value=name)

    _style_header_row(ws, 1, 1 + len(windows))
    _style_header_row(ws, 2, 1 + len(windows))

    # 本体: 各受注 × 各ウィンドウ。セルは "A / B / C" の台数を1文字列で示し、
    # どの工程が動いているかは主工程色で軽く塗る(複数工程が動く場合もある)。
    row = 3
    for order in orders:
        ws.cell(row=row, column=1, value=f"{order} {product_of.get(order, '')}").font = _BOLD
        for wi in range(len(windows)):
            per_stage = [len(counts.get((order, wi, s), set())) for s in _MATRIX_STAGES]
            total = sum(per_stage)
            cell = ws.cell(row=row, column=2 + wi)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER
            if total == 0:
                cell.value = "-"
                continue
            cell.value = "/".join(str(n) for n in per_stage)
            # 最も台数の多い工程の色で塗る(視認性向上)
            top_stage = _MATRIX_STAGES[per_stage.index(max(per_stage))]
            cell.fill = PatternFill("solid", fgColor=_STAGE_FILLS[top_stage])
        row += 1

    # 凡例
    ws.cell(row=row + 1, column=1, value="セルの数値は「工程A / 工程B / 工程C」の稼働号機数").font = Font(italic=True)

    _set_widths(ws, [16] + [7] * len(windows))
    ws.freeze_panes = "B3"


def _add_warnings_sheet(wb: Workbook, result: PlanResult) -> None:
    ws = wb.create_sheet("警告")
    ws.append(["対象", "メッセージ"])
    _style_header_row(ws, 1, 2)
    if not result.warnings:
        ws.append(["-", "警告はありません。"])
    else:
        for w in result.warnings:
            ws.append([w.order_id, w.message])
    _set_widths(ws, [14, 70])
    ws.freeze_panes = "A2"


def export_plan_workbook(result: PlanResult, equipment: EquipmentConfig) -> Workbook:
    """立案結果(PlanResult)をフル内容(稼働率を除く)のExcelワークブックにする。"""
    wb = Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, result)
    _add_schedule_sheet(wb, result)
    _add_matrix_sheet(wb, result, equipment)
    _add_warnings_sheet(wb, result)
    return wb
