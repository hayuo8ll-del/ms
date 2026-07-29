"""現場の **CAP表**(`THM設備Cap/機種対応表`)を読み、設備マスタとして取り込む。

シート構成:
- `THM`     … シフトモード(22h/16h/11h/8h)ごとに、機種×号機の生産可否(○/△/×)、
              号機ごとの設備キャパ、機種別キャパ が並んだブロック
- `機種一覧` … モデル名(RCコード) → 呼称 / 工程別レシピコード の対応

ここから取れるもの:
- `parse_machine_master`  … {シフトモード: [MachineSlot]} 号機ごとの工程/日次能力/生産可否
- `parse_product_aliases` … {RCコード: 呼称} (config の productAliases を実データで置き換える)
- `parse_recipe_codes`    … {レシピコード: 呼称} (TA1_投入計画の機種コード列がこれ)
- `parse_product_daily_caps` … {シフトモード: {呼称: 機種別キャパ}}

機種別キャパは「その機種が使える号機の能力合計」を工程横断で見た最小値にほぼ一致する
(例 Lite-S 42,240 = MIL#3 の1台ぶん)。ただし表に載っていない号機(HAL#1/#2 など)を
使う機種もあるため、**表の 機種別キャパ 列を正**として扱い、号機からの導出は
`derive_product_caps` で参考値として出すだけにする。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import BinaryIO

from openpyxl import load_workbook

# 号機名から工程を引く。「アンテナ検査機」は ANT の検査工程で、投入配分の対象ではない。
_STAGE_PATTERNS = (
    ("ANT", re.compile(r"^ANT")),
    ("TAL", re.compile(r"^TAL")),
    ("HAL", re.compile(r"^HAL")),
    ("MIL", re.compile(r"^MIL")),
)
_MODE_HEADER = re.compile(r"■\s*(\d+)\s*h")
_OK, _COND = "○", "△"

# 現在稼働している号機。CAP表には過去の号機や、シフトモードごとにラベルが食い違う列
# (22Hだけ `ANT#1`、16H以下は `TAL#1`)が残っているため、ここで明示的に絞り込む。
# ANT工程は号機を持たない(工程展開のリードタイムだけで表現する)。
ACTIVE_MACHINES = frozenset({
    "TAL#2", "TAL#3",
    "HAL#5", "HAL#6", "HAL#7", "HAL#8", "HAL#9",
    "MIL#3", "MIL#5", "MIL#6", "MIL#7",
})


@dataclass
class MachineSlot:
    """ある号機の、あるシフトモードでの能力と生産可能機種。"""

    machine_id: str
    stage_id: str
    daily_capacity: float
    # 呼称 -> "○"(生産可) / "△"(条件付き)。× は落とすので入らない。
    eligibility: dict[str, str] = field(default_factory=dict)

    def can_run(self, product: str, allow_conditional: bool = True) -> bool:
        mark = self.eligibility.get(product)
        if mark is None:
            return False
        return mark == _OK or (allow_conditional and mark == _COND)


def _stage_of(machine_name: str) -> str | None:
    for stage_id, pattern in _STAGE_PATTERNS:
        if pattern.match(machine_name):
            return stage_id
    return None


def _norm(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _mark(value: object) -> str | None:
    """セルの ○/〇/△/× を正規化する(全角の〇と○が混在している)。"""
    text = _norm(value)
    if text in ("○", "〇", "◯"):
        return _OK
    if text in ("△", "▲"):
        return _COND
    return None  # × / 空欄 = 生産不可


def _mode_blocks(ws) -> list[tuple[str, int]]:
    """(シフトモード名, ヘッダー行) の一覧を返す。ヘッダー行に号機名が並ぶ。"""
    blocks: list[tuple[str, int]] = []
    for r in range(1, ws.max_row + 1):
        m = _MODE_HEADER.match(_norm(ws.cell(row=r, column=1).value))
        if m:
            blocks.append((f"{int(m.group(1))}h", r + 1))
    return blocks


def parse_machine_master(
    file_obj: BinaryIO,
    only_machines: set[str] | None = None,
) -> dict[str, list[MachineSlot]]:
    """CAP表の `THM` シートから {シフトモード: [MachineSlot]} を読む。

    `only_machines` を渡すと、その号機だけを採る。CAP表には過去の号機や、
    シフトモードごとにラベルが食い違う列(22Hだけ `ANT#1`、16H以下は `TAL#1`)が
    残っているため、**実際に稼働している号機**を明示して絞り込むために使う。
    """
    ws = load_workbook(file_obj, data_only=True)["THM"]
    result: dict[str, list[MachineSlot]] = {}

    for mode, header_row in _mode_blocks(ws):
        # ヘッダー行の号機名(工程が引ける列だけ採用。「アンテナ検査機」等は対象外)
        columns: dict[int, str] = {}
        for c in range(5, ws.max_column + 1):
            name = _norm(ws.cell(row=header_row, column=c).value)
            if not name or not _stage_of(name):
                continue
            if only_machines is not None and name not in only_machines:
                continue
            columns[c] = name
        if not columns:
            continue

        # 機種行: ヘッダーの2行下から「設備キャパ」行の手前まで
        product_rows: list[tuple[int, str]] = []
        capacity_row = None
        for r in range(header_row + 2, ws.max_row + 1):
            if _norm(ws.cell(row=r, column=4).value) == "設備キャパ":
                capacity_row = r
                break
            name = _norm(ws.cell(row=r, column=1).value)
            if name:
                product_rows.append((r, name))
        if capacity_row is None:
            continue

        slots: list[MachineSlot] = []
        for c, machine_id in columns.items():
            cap = ws.cell(row=capacity_row, column=c).value
            slot = MachineSlot(
                machine_id=machine_id,
                stage_id=_stage_of(machine_id) or "",
                daily_capacity=float(cap) if isinstance(cap, (int, float)) else 0.0,
            )
            for r, product in product_rows:
                mark = _mark(ws.cell(row=r, column=c).value)
                if mark:
                    slot.eligibility[product] = mark
            slots.append(slot)
        result[mode] = slots
    return result


def parse_product_daily_caps(file_obj: BinaryIO) -> dict[str, dict[str, float]]:
    """CAP表の 機種別キャパ 列から {シフトモード: {呼称: 日産キャパ}} を読む。

    キャパ列が数値でない機種(Suica4 の `×` など)は生産不可として入れない。
    """
    ws = load_workbook(file_obj, data_only=True)["THM"]
    result: dict[str, dict[str, float]] = {}

    for mode, header_row in _mode_blocks(ws):
        cap_col = None
        for c in range(5, ws.max_column + 1):
            if "機種別" in _norm(ws.cell(row=header_row, column=c).value):
                cap_col = c
                break
        if cap_col is None:
            continue
        caps: dict[str, float] = {}
        for r in range(header_row + 2, ws.max_row + 1):
            if _norm(ws.cell(row=r, column=4).value) == "設備キャパ":
                break
            product = _norm(ws.cell(row=r, column=1).value)
            value = ws.cell(row=r, column=cap_col).value
            if product and isinstance(value, (int, float)) and value:
                caps[product] = float(value)
        result[mode] = caps
    return result


def parse_product_aliases(file_obj: BinaryIO) -> dict[str, str]:
    """`機種一覧` から {モデル名(RCコード): 呼称} を読む。

    `resolve_product` は**最長前方一致**なので、枝番付きのフルのモデル名だけを鍵に
    すると `RC-S103/JW16　INLAY` のような表記に対して台帳側の `RC-S103/JW16` が
    引けなくなる(鍵のほうが長いと前方一致しない)。そこで枝番を落とした
    **ルートコード**(`RC-S103`)も併せて登録する。ただしルートが複数の呼称に割れる
    場合(1つのRC番号が別機種に跨る場合)は曖昧なので登録しない。
    """
    ws = load_workbook(file_obj, data_only=True)["機種一覧"]
    aliases: dict[str, str] = {}
    roots: dict[str, set[str]] = {}

    for r in range(2, ws.max_row + 1):
        product = _norm(ws.cell(row=r, column=1).value)
        model = _norm(ws.cell(row=r, column=2).value)
        if not product or not model or model == "-":
            continue
        aliases[model] = product
        roots.setdefault(model.split("/")[0].strip(), set()).add(product)

    for root, products in roots.items():
        if len(products) == 1 and root not in aliases:
            aliases[root] = next(iter(products))
    return aliases


def parse_recipe_codes(file_obj: BinaryIO) -> dict[str, str]:
    """`機種一覧` の工程別レシピコード → 呼称。TA1_投入計画の機種コード列がこれ。"""
    ws = load_workbook(file_obj, data_only=True)["機種一覧"]
    codes: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        product = _norm(ws.cell(row=r, column=1).value)
        if not product:
            continue
        for c in (3, 4, 5):  # TAL/HAL/MILレシピ
            code = _norm(ws.cell(row=r, column=c).value)
            if code and code != "-":
                codes.setdefault(code.split("/")[0], product)
    return codes


def derive_product_caps(
    machines: list[MachineSlot], allow_conditional: bool = False
) -> dict[str, float]:
    """号機マスタから機種別キャパを導出する(表の値との突き合わせ用の参考値)。

    その機種が使える号機の能力を工程ごとに合計し、**工程間の最小値**を取る
    (直列ラインなので一番細い工程が律速)。
    """
    by_stage: dict[str, dict[str, float]] = {}
    for slot in machines:
        for product, mark in slot.eligibility.items():
            if mark == _COND and not allow_conditional:
                continue
            by_stage.setdefault(product, {})
            by_stage[product][slot.stage_id] = (
                by_stage[product].get(slot.stage_id, 0.0) + slot.daily_capacity
            )
    return {
        product: min(stages.values())
        for product, stages in by_stage.items()
        if stages and min(stages.values()) > 0
    }
