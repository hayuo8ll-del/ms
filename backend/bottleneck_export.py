"""ボトルネック計画(bottleneck_planner)の結果を、現場の2つの表の形でExcel出力する。

  1. 生産計画(機種×日): 機種ごとに工程(ANT/TAL/HAL/MIL)の日次台数を並べる
     (TA1_生産計画 の形)。
  2. 製番別MIL: 出荷ロット(製番)ごとのMIL完成日・納期・判定
     (THM 短期投入予定表 の形)。

加えて サマリー / 警告 シートを持つ。
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bottleneck_planner import BottleneckPlanResult, DemandItem

_HEADER_FILL = PatternFill("solid", fgColor="2D3139")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_STAGE_FILL = PatternFill("solid", fgColor="EEF1F5")
_MIL_FILL = PatternFill("solid", fgColor="FCE4C6")  # オレンジ(現場のMIL塗り)
_LATE_FILL = PatternFill("solid", fgColor="F8D7D2")  # 納期超過(赤系)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="D0D4DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_DATE_FMT = "m/d"


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_summary(
    wb: Workbook,
    result: BottleneckPlanResult,
    demands: list[DemandItem],
    extra_summary: list[tuple[str, object]] | None = None,
) -> None:
    ws = wb.create_sheet("サマリー")
    total_qty = sum(d.quantity for d in demands)
    days = result.working_days
    rows = [
        ("計画期間", f"{days[0].isoformat()} 〜 {days[-1].isoformat()}" if days else "-"),
        ("稼働日数", len(days)),
        ("選択シフト", result.shift_mode),
        ("日次能力(ボトルネック)", result.daily_capacity),
        ("必要日次レート", round(result.required_daily_rate)),
        ("対象ロット数(製番)", len(demands)),
        ("需要総数(残数量)", total_qty),
        ("MIL納期超過ロット数", sum(1 for lot in result.mil_lots if lot.on_time is False)),
        ("警告件数", len(result.warnings)),
    ]
    if extra_summary:
        rows.extend(extra_summary)
    ws.append(["項目", "値"])
    _style_header(ws, 1, 2)
    for label, value in rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = _BOLD
    _set_widths(ws, [24, 30])
    ws.freeze_panes = "A2"


def _add_stage_matrix(wb: Workbook, result: BottleneckPlanResult, stage_order: list[str]) -> None:
    """機種×工程 × 日 の日次台数マトリクス(TA1_生産計画 の形)。"""
    ws = wb.create_sheet("生産計画(機種×日)")
    days = result.working_days

    # (product, stage, day) -> qty
    agg: dict[tuple[str, str, date], float] = {}
    products_first_day: dict[str, int] = {}
    day_index = {d: i for i, d in enumerate(days)}
    for c in result.stage_allocation:
        agg[(c.product, c.stage_id, c.day)] = agg.get((c.product, c.stage_id, c.day), 0.0) + c.quantity
        if c.stage_id == (stage_order[0] if stage_order else c.stage_id):
            pass
    # 機種の並び順: HAL(ボトルネック)の初日が早い順。無ければ任意。
    bottleneck = "HAL" if any(s == "HAL" for s in stage_order) else (stage_order[-1] if stage_order else None)
    for c in result.stage_allocation:
        if c.stage_id == bottleneck:
            i = day_index.get(c.day, 10**9)
            products_first_day[c.product] = min(products_first_day.get(c.product, 10**9), i)
    products = sorted(products_first_day, key=lambda p: (products_first_day[p], p))
    # stage_allocation に現れる全機種を取りこぼさない
    for c in result.stage_allocation:
        if c.product not in products:
            products.append(c.product)

    # ヘッダー: 機種 | 工程 | 日付...
    header = ["機種", "工程"] + [d.strftime("%-m/%-d") for d in days]
    ws.append(header)
    _style_header(ws, 1, len(header))

    for product in products:
        for stage in stage_order:
            row = [product if stage == stage_order[0] else "", stage]
            for d in days:
                q = agg.get((product, stage, d), 0.0)
                row.append(q if q else None)
            ws.append(row)
            r = ws.max_row
            ws.cell(row=r, column=1).font = _BOLD
            ws.cell(row=r, column=2).fill = _STAGE_FILL
            ws.cell(row=r, column=2).font = _BOLD

    _set_widths(ws, [16, 8] + [7] * len(days))
    ws.freeze_panes = "C2"


def _add_mil_lots(wb: Workbook, result: BottleneckPlanResult) -> None:
    """製番(出荷ロット)別のMIL完成予定(THM 短期投入予定表 の形)。"""
    ws = wb.create_sheet("製番別MIL")
    headers = ["製番", "機種", "数量", "MIL完成予定", "納期", "判定"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for lot in result.mil_lots:
        due = lot.due_date.isoformat() if lot.due_date else "-"
        if lot.on_time is None:
            judge = "-"
        else:
            judge = "○ 納期内" if lot.on_time else "× 超過"
        ws.append([lot.order_id, lot.product, lot.quantity, lot.completion_day, due, judge])
        r = ws.max_row
        ws.cell(row=r, column=4).number_format = _DATE_FMT
        # MILはオレンジ、納期超過行は赤で強調
        fill = _LATE_FILL if lot.on_time is False else _MIL_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _BORDER
        ws.cell(row=r, column=6).fill = fill

    _set_widths(ws, [12, 14, 10, 13, 12, 10])
    ws.freeze_panes = "A2"


def _add_progress(wb: Workbook, result: BottleneckPlanResult) -> None:
    """計画/実績/差/累計の進捗(現場のSheet1の形)。列=稼働日、行=各指標。"""
    if not result.progress:
        return
    ws = wb.create_sheet("進捗")
    days = [p.day for p in result.progress]
    has_actual = any(p.actual is not None for p in result.progress)

    header = ["指標"] + [d.strftime("%-m/%-d") for d in days]
    ws.append(header)
    _style_header(ws, 1, len(header))

    def add_row(label: str, values: list[object]) -> None:
        ws.append([label] + values)
        ws.cell(row=ws.max_row, column=1).font = _BOLD

    add_row("計画", [p.plan or None for p in result.progress])
    add_row("計画累計", [p.plan_cum for p in result.progress])
    if has_actual:
        add_row("実績", [p.actual for p in result.progress])
        add_row("実績累計", [p.actual_cum for p in result.progress])
        add_row("差(実績-計画)", [p.diff for p in result.progress])
        add_row("進捗(累計)", [p.progress_cum for p in result.progress])
        # 進捗が負(遅れ)のセルを赤く
        prog_row = ws.max_row
        for i, p in enumerate(result.progress):
            if p.progress_cum is not None and p.progress_cum < 0:
                ws.cell(row=prog_row, column=2 + i).fill = _LATE_FILL

    _set_widths(ws, [14] + [7] * len(days))
    ws.freeze_panes = "B2"


def _add_warnings(wb: Workbook, result: BottleneckPlanResult) -> None:
    ws = wb.create_sheet("警告")
    ws.append(["メッセージ"])
    _style_header(ws, 1, 1)
    if not result.warnings:
        ws.append(["警告はありません。"])
    else:
        for w in result.warnings:
            ws.append([w])
    _set_widths(ws, [110])
    ws.freeze_panes = "A2"


def export_bottleneck_workbook(
    result: BottleneckPlanResult,
    demands: list[DemandItem],
    stage_order: list[str],
    extra_summary: list[tuple[str, object]] | None = None,
) -> Workbook:
    """ボトルネック計画結果を4シートのワークブックにする。"""
    wb = Workbook()
    wb.remove(wb.active)
    _add_summary(wb, result, demands, extra_summary=extra_summary)
    _add_stage_matrix(wb, result, stage_order)
    _add_mil_lots(wb, result)
    _add_progress(wb, result)
    _add_warnings(wb, result)
    return wb
