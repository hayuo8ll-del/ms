"""config/*.json とExcelワークブックを相互変換する(実データ取り込み用)。

Web画面の「Excelを取り込む」機能で使う。1つの.xlsxファイルに以下のシートを含める
(シート名・列名は固定。`export_workbook()` で現在のconfig内容を埋めたテンプレートを
ダウンロードできる):

  Stages / Machines / LotSplitting / ShiftModes / Settings
  Changeover / AShiftOnlyTransitions
  Orders / Inventory / RawMaterials / RawMaterialIncoming

各シートは1行目がヘッダー、2行目以降が1行1レコードの表形式。
取り込みに失敗した場合は元の config/*.json は一切変更しない(全件検証してから書き込む)。
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"

REQUIRED_SHEETS = [
    "Stages",
    "Machines",
    "LotSplitting",
    "ShiftModes",
    "Settings",
    "Changeover",
    "Orders",
    "Inventory",
    "RawMaterials",
]
OPTIONAL_SHEETS = ["AShiftOnlyTransitions", "RawMaterialIncoming"]


@dataclass
class ImportIssue:
    sheet: str
    row: int | None
    message: str


class ImportValidationError(Exception):
    def __init__(self, issues: list[ImportIssue]):
        self.issues = issues
        super().__init__(f"{len(issues)}件の入力エラーがあります。")


class WorkbookReadError(Exception):
    pass


@dataclass
class ImportSummary:
    stages: int
    machines: int
    changeover_rules: int
    orders: int
    inventory_items: int
    raw_materials: int


# -- セル値の正規化 -----------------------------------------------------------------


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _cell_str(value).upper() in ("TRUE", "1", "YES", "Y")


def _cell_time(value: Any) -> str:
    """Excelのtime/datetimeセル、または \"08:30\" 文字列を \"HH:MM\" に正規化する。"""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return _cell_str(value)


def _cell_date_str(value: Any) -> str:
    """Excelのdate/datetimeセル、または \"2026-07-25\" 文字列を \"YYYY-MM-DD\" に正規化する。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _cell_str(value)


def _cell_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# -- シート読み込み -----------------------------------------------------------------


def _read_rows(ws: Worksheet) -> list[tuple[int, dict[str, Any]]]:
    """1行目をヘッダーとして、2行目以降を(行番号, {ヘッダー名: 値})で返す。空行はスキップ。"""
    rows = list(ws.iter_rows())
    if not rows:
        return []
    headers = [_cell_str(c.value) for c in rows[0]]
    result: list[tuple[int, dict[str, Any]]] = []
    for r in rows[1:]:
        values = {headers[i]: r[i].value for i in range(len(headers)) if headers[i]}
        if all(v is None or v == "" for v in values.values()):
            continue
        result.append((r[0].row, values))
    return result


def parse_workbook(file_obj: BinaryIO) -> tuple[dict, dict, dict, ImportSummary]:
    """アップロードされたワークブックを検証し、equipment/changeover/orders の
    JSON互換dictを構築する。検証エラーがあれば ImportValidationError を送出する
    (この場合、呼び出し側は config/*.json を書き換えてはならない)。
    """
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001 - 壊れたファイル/非対応形式を一律エラーにする
        raise WorkbookReadError("Excelファイル(.xlsx)として読み込めませんでした。") from exc

    issues: list[ImportIssue] = []

    def sheet_rows(name: str, required: bool) -> list[tuple[int, dict[str, Any]]]:
        if name not in wb.sheetnames:
            if required:
                issues.append(ImportIssue(name, None, f"シート「{name}」が見つかりません。"))
            return []
        return _read_rows(wb[name])

    stage_rows = sheet_rows("Stages", True)
    machine_rows = sheet_rows("Machines", True)
    lot_split_rows = sheet_rows("LotSplitting", True)
    shift_rows = sheet_rows("ShiftModes", True)
    settings_rows = sheet_rows("Settings", True)
    changeover_rows = sheet_rows("Changeover", True)
    a_shift_rows = sheet_rows("AShiftOnlyTransitions", False)
    order_rows = sheet_rows("Orders", True)
    inventory_rows = sheet_rows("Inventory", True)
    raw_material_rows = sheet_rows("RawMaterials", True)
    incoming_rows = sheet_rows("RawMaterialIncoming", False)

    # --- Stages ---
    stages_by_id: dict[str, dict] = {}
    for row_no, row in stage_rows:
        stage_id = _cell_str(row.get("stageId"))
        if not stage_id:
            issues.append(ImportIssue("Stages", row_no, "stageId が空です。"))
            continue
        if stage_id in stages_by_id:
            issues.append(ImportIssue("Stages", row_no, f"stageId「{stage_id}」が重複しています。"))
            continue
        order_val = _cell_number(row.get("order"))
        if order_val is None:
            issues.append(ImportIssue("Stages", row_no, "order は数値で指定してください。"))
            continue
        stage: dict[str, Any] = {
            "stageId": stage_id,
            "stageName": _cell_str(row.get("stageName")) or stage_id,
            "order": int(order_val),
            "machines": [],
            "uninterruptible": _cell_bool(row.get("uninterruptible")),
        }
        batch_rounding_raw = row.get("batchRounding")
        if batch_rounding_raw not in (None, ""):
            batch_rounding = _cell_number(batch_rounding_raw)
            if batch_rounding is None:
                issues.append(ImportIssue("Stages", row_no, "batchRounding は数値で指定してください。"))
            else:
                stage["batchRounding"] = int(batch_rounding)
        stages_by_id[stage_id] = stage

    if not stages_by_id:
        issues.append(ImportIssue("Stages", None, "少なくとも1つの工程が必要です。"))

    # --- Machines ---
    machine_ids_seen: set[str] = set()
    for row_no, row in machine_rows:
        stage_id = _cell_str(row.get("stageId"))
        machine_id = _cell_str(row.get("machineId"))
        if stage_id not in stages_by_id:
            issues.append(ImportIssue("Machines", row_no, f"未定義の stageId「{stage_id}」です。"))
            continue
        if not machine_id:
            issues.append(ImportIssue("Machines", row_no, "machineId が空です。"))
            continue
        if machine_id in machine_ids_seen:
            issues.append(ImportIssue("Machines", row_no, f"machineId「{machine_id}」が重複しています。"))
            continue
        capacity = _cell_number(row.get("capacityPerHour"))
        if capacity is None or capacity <= 0:
            issues.append(ImportIssue("Machines", row_no, "capacityPerHour は正の数値で指定してください。"))
            continue
        machine_ids_seen.add(machine_id)
        stages_by_id[stage_id]["machines"].append(
            {"machineId": machine_id, "name": _cell_str(row.get("name")) or machine_id, "capacityPerHour": capacity}
        )

    for stage_id, stage in stages_by_id.items():
        if not stage["machines"]:
            issues.append(ImportIssue("Machines", None, f"工程「{stage_id}」に号機が1台も定義されていません。"))

    # --- LotSplitting ---
    lot_split_after: str | None = None
    lot_split_into = 1
    if not lot_split_rows:
        issues.append(ImportIssue("LotSplitting", None, "1行(stageAfter, splitInto)を入力してください。"))
    else:
        row_no, row = lot_split_rows[0]
        lot_split_after = _cell_str(row.get("stageAfter")) or None
        if lot_split_after and lot_split_after not in stages_by_id:
            issues.append(ImportIssue("LotSplitting", row_no, f"未定義の stageAfter「{lot_split_after}」です。"))
        split_into_val = _cell_number(row.get("splitInto"))
        if split_into_val is None or split_into_val < 1:
            issues.append(ImportIssue("LotSplitting", row_no, "splitInto は1以上の数値で指定してください。"))
        else:
            lot_split_into = int(split_into_val)

    # --- ShiftModes ---
    shift_modes: dict[str, list[dict]] = {}
    for row_no, row in shift_rows:
        mode = _cell_str(row.get("shiftMode"))
        shift_name = _cell_str(row.get("shiftName"))
        start = _cell_time(row.get("start"))
        end = _cell_time(row.get("end"))
        if not mode or not shift_name or not start or not end:
            issues.append(ImportIssue("ShiftModes", row_no, "shiftMode/shiftName/start/end はすべて必須です。"))
            continue
        shift_modes.setdefault(mode, []).append({"shiftName": shift_name, "start": start, "end": end})

    if not shift_modes:
        issues.append(ImportIssue("ShiftModes", None, "少なくとも1つのシフトパターンが必要です。"))

    # --- Settings ---
    settings = {_cell_str(row.get("key")): row.get("value") for _row_no, row in settings_rows}
    default_shift_mode = _cell_str(settings.get("defaultShiftMode"))
    if not default_shift_mode:
        issues.append(ImportIssue("Settings", None, "key=defaultShiftMode の行が必要です。"))
    elif default_shift_mode not in shift_modes:
        issues.append(
            ImportIssue("Settings", None, f"defaultShiftMode「{default_shift_mode}」がShiftModesシートに定義されていません。")
        )

    plan_start_raw = settings.get("planStart")
    plan_start: str | None = None
    if isinstance(plan_start_raw, datetime):
        plan_start = plan_start_raw.isoformat()
    else:
        plan_start_str = _cell_str(plan_start_raw)
        if not plan_start_str:
            issues.append(ImportIssue("Settings", None, "key=planStart の行が必要です。"))
        else:
            try:
                datetime.fromisoformat(plan_start_str)
                plan_start = plan_start_str
            except ValueError:
                issues.append(ImportIssue("Settings", None, "planStart はISO日時形式(例: 2026-07-19T08:30:00)で指定してください。"))

    # --- Changeover ---
    changeover: dict[str, dict[str, dict[str, float]]] = {}
    products_seen: set[str] = set()
    changeover_count = 0
    for row_no, row in changeover_rows:
        stage_id = _cell_str(row.get("stageId"))
        from_p = _cell_str(row.get("fromProduct"))
        to_p = _cell_str(row.get("toProduct"))
        if stage_id not in stages_by_id:
            issues.append(ImportIssue("Changeover", row_no, f"未定義の stageId「{stage_id}」です。"))
            continue
        if not from_p or not to_p:
            issues.append(ImportIssue("Changeover", row_no, "fromProduct/toProduct は必須です。"))
            continue
        minutes = _cell_number(row.get("minutes"))
        if minutes is None or minutes < 0:
            issues.append(ImportIssue("Changeover", row_no, "minutes は0以上の数値で指定してください。"))
            continue
        changeover.setdefault(stage_id, {}).setdefault(from_p, {})[to_p] = minutes
        products_seen.update([from_p, to_p])
        changeover_count += 1

    # --- AShiftOnlyTransitions ---
    a_shift_only: dict[str, list[dict[str, str]]] = {}
    for row_no, row in a_shift_rows:
        stage_id = _cell_str(row.get("stageId"))
        from_p = _cell_str(row.get("fromProduct"))
        to_p = _cell_str(row.get("toProduct"))
        if stage_id not in stages_by_id:
            issues.append(ImportIssue("AShiftOnlyTransitions", row_no, f"未定義の stageId「{stage_id}」です。"))
            continue
        if not from_p or not to_p:
            issues.append(ImportIssue("AShiftOnlyTransitions", row_no, "fromProduct/toProduct は必須です。"))
            continue
        a_shift_only.setdefault(stage_id, []).append({"from": from_p, "to": to_p})

    # --- Orders ---
    orders: list[dict] = []
    order_ids_seen: set[str] = set()
    for row_no, row in order_rows:
        order_id = _cell_str(row.get("orderId"))
        product = _cell_str(row.get("product"))
        if not order_id:
            issues.append(ImportIssue("Orders", row_no, "orderId が空です。"))
            continue
        if order_id in order_ids_seen:
            issues.append(ImportIssue("Orders", row_no, f"orderId「{order_id}」が重複しています。"))
            continue
        order_ids_seen.add(order_id)
        if not product:
            issues.append(ImportIssue("Orders", row_no, "product が空です。"))
            continue
        quantity = _cell_number(row.get("quantity"))
        if quantity is None or quantity <= 0:
            issues.append(ImportIssue("Orders", row_no, "quantity は正の数値で指定してください。"))
            continue
        due_date = _cell_date_str(row.get("dueDate"))
        try:
            date.fromisoformat(due_date)
        except ValueError:
            issues.append(ImportIssue("Orders", row_no, "dueDate は日付形式(例: 2026-07-25)で指定してください。"))
            continue
        orders.append({"orderId": order_id, "product": product, "quantity": quantity, "dueDate": due_date})
        products_seen.add(product)

    if not orders:
        issues.append(ImportIssue("Orders", None, "少なくとも1件の受注が必要です。"))

    # --- Inventory ---
    inventory: dict[str, dict] = {}
    for row_no, row in inventory_rows:
        product = _cell_str(row.get("product"))
        if not product:
            issues.append(ImportIssue("Inventory", row_no, "product が空です。"))
            continue
        current = _cell_number(row.get("currentStock"))
        safety = _cell_number(row.get("safetyStock"))
        if current is None or safety is None:
            issues.append(ImportIssue("Inventory", row_no, "currentStock/safetyStock は数値で指定してください。"))
            continue
        inventory[product] = {"currentStock": current, "safetyStock": safety}

    # --- RawMaterials / RawMaterialIncoming ---
    raw_materials: dict[str, dict] = {}
    for row_no, row in raw_material_rows:
        product = _cell_str(row.get("product"))
        material_id = _cell_str(row.get("materialId"))
        if not product or not material_id:
            issues.append(ImportIssue("RawMaterials", row_no, "materialId/product は必須です。"))
            continue
        on_hand = _cell_number(row.get("onHand"))
        if on_hand is None or on_hand < 0:
            issues.append(ImportIssue("RawMaterials", row_no, "onHand は0以上の数値で指定してください。"))
            continue
        raw_materials[product] = {"materialId": material_id, "onHand": on_hand, "incoming": []}

    for row_no, row in incoming_rows:
        product = _cell_str(row.get("product"))
        if product not in raw_materials:
            issues.append(
                ImportIssue("RawMaterialIncoming", row_no, f"未定義の product「{product}」です(先にRawMaterialsシートへ追加してください)。")
            )
            continue
        inc_date = _cell_date_str(row.get("date"))
        try:
            date.fromisoformat(inc_date)
        except ValueError:
            issues.append(ImportIssue("RawMaterialIncoming", row_no, "date は日付形式で指定してください。"))
            continue
        inc_qty = _cell_number(row.get("quantity"))
        if inc_qty is None or inc_qty <= 0:
            issues.append(ImportIssue("RawMaterialIncoming", row_no, "quantity は正の数値で指定してください。"))
            continue
        raw_materials[product]["incoming"].append({"date": inc_date, "quantity": inc_qty})

    if issues:
        raise ImportValidationError(issues)

    equipment = {
        "stages": sorted(stages_by_id.values(), key=lambda s: s["order"]),
        "lotSplitting": {"stageAfter": lot_split_after, "splitInto": lot_split_into},
        "shiftModes": shift_modes,
        "defaultShiftMode": default_shift_mode,
    }
    changeover_doc = {
        "products": sorted(products_seen),
        **changeover,
        "aShiftOnlyTransitions": a_shift_only,
    }
    orders_doc = {
        "orders": orders,
        "inventory": inventory,
        "rawMaterials": raw_materials,
        "planStart": plan_start,
    }
    summary = ImportSummary(
        stages=len(stages_by_id),
        machines=len(machine_ids_seen),
        changeover_rules=changeover_count,
        orders=len(orders),
        inventory_items=len(inventory),
        raw_materials=len(raw_materials),
    )
    return equipment, changeover_doc, orders_doc, summary


# -- config/*.json への保存 -----------------------------------------------------------


def save_config(equipment: dict, changeover: dict, orders_doc: dict) -> None:
    _write_json("equipment_master.json", equipment)
    _write_json("changeover_matrix.json", changeover)
    _write_json("orders_sample.json", orders_doc)


def _write_json(name: str, data: dict) -> None:
    with open(CONFIG_DIR / name, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


# -- テンプレート/現状データのExcel出力 -------------------------------------------------


def _load_raw_json(name: str) -> dict:
    with open(CONFIG_DIR / name, encoding="utf-8") as f:
        return json.load(f)


def _write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def export_workbook() -> Workbook:
    """現在の config/*.json の内容を埋め込んだExcelワークブックを生成する。

    そのままダウンロードして値を書き換え、再アップロードすれば取り込める
    「取り込みテンプレート」として使う。
    """
    equipment_raw = _load_raw_json("equipment_master.json")
    changeover_raw = _load_raw_json("changeover_matrix.json")
    orders_raw = _load_raw_json("orders_sample.json")

    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "Stages",
        ["stageId", "stageName", "order", "uninterruptible", "batchRounding"],
        [
            [s["stageId"], s["stageName"], s["order"], s.get("uninterruptible", False), s.get("batchRounding", "")]
            for s in equipment_raw["stages"]
        ],
    )
    _write_sheet(
        wb,
        "Machines",
        ["stageId", "machineId", "name", "capacityPerHour"],
        [
            [s["stageId"], m["machineId"], m["name"], m["capacityPerHour"]]
            for s in equipment_raw["stages"]
            for m in s["machines"]
        ],
    )
    lot = equipment_raw.get("lotSplitting", {})
    _write_sheet(wb, "LotSplitting", ["stageAfter", "splitInto"], [[lot.get("stageAfter", ""), lot.get("splitInto", 1)]])

    shift_rows = [
        [mode, d["shiftName"], d["start"], d["end"]]
        for mode, defs in equipment_raw.get("shiftModes", {}).items()
        for d in defs
    ]
    _write_sheet(wb, "ShiftModes", ["shiftMode", "shiftName", "start", "end"], shift_rows)

    _write_sheet(
        wb,
        "Settings",
        ["key", "value"],
        [
            ["defaultShiftMode", equipment_raw.get("defaultShiftMode", "")],
            ["planStart", orders_raw.get("planStart", "")],
        ],
    )

    changeover_rows = [
        [stage_id, from_p, to_p, minutes]
        for stage_id, table in changeover_raw.items()
        if stage_id not in ("products", "aShiftOnlyTransitions") and not stage_id.startswith("_")
        for from_p, tos in table.items()
        for to_p, minutes in tos.items()
    ]
    _write_sheet(wb, "Changeover", ["stageId", "fromProduct", "toProduct", "minutes"], changeover_rows)

    a_shift_rows = [
        [stage_id, t["from"], t["to"]]
        for stage_id, transitions in changeover_raw.get("aShiftOnlyTransitions", {}).items()
        for t in transitions
    ]
    _write_sheet(wb, "AShiftOnlyTransitions", ["stageId", "fromProduct", "toProduct"], a_shift_rows)

    _write_sheet(
        wb,
        "Orders",
        ["orderId", "product", "quantity", "dueDate"],
        [[o["orderId"], o["product"], o["quantity"], o["dueDate"]] for o in orders_raw.get("orders", [])],
    )
    _write_sheet(
        wb,
        "Inventory",
        ["product", "currentStock", "safetyStock"],
        [[p, v["currentStock"], v["safetyStock"]] for p, v in orders_raw.get("inventory", {}).items()],
    )
    _write_sheet(
        wb,
        "RawMaterials",
        ["materialId", "product", "onHand"],
        [[v["materialId"], p, v["onHand"]] for p, v in orders_raw.get("rawMaterials", {}).items()],
    )
    incoming_rows = [
        [p, inc["date"], inc["quantity"]]
        for p, v in orders_raw.get("rawMaterials", {}).items()
        for inc in v.get("incoming", [])
    ]
    _write_sheet(wb, "RawMaterialIncoming", ["product", "date", "quantity"], incoming_rows)

    return wb
