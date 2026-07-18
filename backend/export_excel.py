"""Export schedule_result.json to an Excel workbook (gantt data, warnings, utilization).

Expected schedule_result.json shape::

    {
        "tasks": [
            {
                "process": "組立",
                "equipment": "設備A",
                "item": "品番001",
                "quantity": 100,
                "start": "2026-07-18T08:00:00",
                "end": "2026-07-18T10:00:00"
            },
            ...
        ],
        "warnings": ["設備Aで稼働率が100%を超えています", ...]
    }
"""
import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

GANTT_COLUMNS = ["工程", "設備", "品目", "数量", "開始", "終了"]


def load_data(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def parse_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).replace("Z", "+00:00")
    return datetime.fromisoformat(text)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def autosize_columns(ws, min_width=10, max_width=40):
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(max(length + 2, min_width), max_width)


def build_gantt_sheet(wb, tasks):
    ws = wb.active
    ws.title = "ガントデータ"
    ws.append(GANTT_COLUMNS)
    for task in tasks:
        ws.append([
            task.get("process", ""),
            task.get("equipment", ""),
            task.get("item", ""),
            task.get("quantity", ""),
            task.get("start", ""),
            task.get("end", ""),
        ])
    style_header_row(ws)
    ws.freeze_panes = "A2"
    autosize_columns(ws)
    return ws


def build_warnings_sheet(wb, warnings):
    ws = wb.create_sheet("警告")
    ws.append(["No.", "警告内容"])
    for i, warning in enumerate(warnings, start=1):
        ws.append([i, warning])
    style_header_row(ws)
    if not warnings:
        ws.append(["-", "警告はありません"])
    autosize_columns(ws)
    return ws


def compute_utilization(tasks):
    """Return list of (equipment, busy_hours, utilization_rate_percent)."""
    spans = {}
    starts = []
    ends = []
    for task in tasks:
        start = parse_datetime(task.get("start"))
        end = parse_datetime(task.get("end"))
        equipment = task.get("equipment", "不明")
        if start is None or end is None:
            continue
        starts.append(start)
        ends.append(end)
        spans.setdefault(equipment, 0.0)
        spans[equipment] += (end - start).total_seconds() / 3600.0

    if not starts:
        return []

    period_hours = (max(ends) - min(starts)).total_seconds() / 3600.0
    if period_hours <= 0:
        period_hours = 1.0

    results = []
    for equipment, busy_hours in sorted(spans.items()):
        rate = min(busy_hours / period_hours * 100.0, 100.0)
        results.append((equipment, round(busy_hours, 2), round(rate, 1)))
    return results


def build_utilization_sheet(wb, tasks):
    ws = wb.create_sheet("稼働率")
    ws.append(["設備", "稼働時間(h)", "稼働率(%)"])
    utilization = compute_utilization(tasks)
    for equipment, busy_hours, rate in utilization:
        ws.append([equipment, busy_hours, rate])
    style_header_row(ws)
    autosize_columns(ws)

    if utilization:
        chart = BarChart()
        chart.type = "col"
        chart.title = "設備別稼働率"
        chart.y_axis.title = "稼働率(%)"
        chart.x_axis.title = "設備"
        last_row = len(utilization) + 1
        data = Reference(ws, min_col=3, min_row=1, max_row=last_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "E2")

    return ws


def build_workbook(data):
    wb = Workbook()
    tasks = data.get("tasks", [])
    warnings = data.get("warnings", [])
    build_gantt_sheet(wb, tasks)
    build_warnings_sheet(wb, warnings)
    build_utilization_sheet(wb, tasks)
    return wb


def main():
    parser = argparse.ArgumentParser(description="Export schedule_result.json to Excel")
    parser.add_argument("--input", default="schedule_result.json", help="Input JSON path")
    parser.add_argument("--output", default="schedule_result.xlsx", help="Output Excel path")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"入力ファイルが見つかりません: {input_path}")

    data = load_data(input_path)
    wb = build_workbook(data)
    wb.save(args.output)
    print(f"Excelファイルを出力しました: {args.output}")


if __name__ == "__main__":
    main()
