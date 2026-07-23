"""THM 生産台帳(受注/納期の実データ)を、ボトルネック計画エンジンの需要入力へ変換する。

台帳シートは1行目が空、2行目がヘッダー、3行目以降が1行1受注:
  № / ライン / 完成品名 / 完成品コード / 製番 / ICチップ / … /
  着手予定日 / 完成予定日 / 完成予定数 / …

出荷ロットの識別子(order_id)には **製番列** を使う(現場のMIL表・投入予定表と同じキー。
製番が空の行のみ№で代用)。

機種(呼称)は、ICチップ列(「東芝さそり」等は金融/交通を区別できない)ではなく、
**完成品コード/完成品名(RC-コード)→呼称** の対応で解決する(最長一致)。既定の対応表
`PRODUCT_ALIASES` は機種一覧(CAP)から生成したもの。会社データでずれる場合は差し替え可。

同じワークブックに「実績」シート(列: 製番 / 実績数)を足しておくと、`parse_actuals` で
製番別の生産実績を読み取れる(計画側で残数量に控除して再立案するのに使う)。
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import BinaryIO

from openpyxl import load_workbook

from bottleneck_planner import DemandItem

# 機種別キャパ(日産台数)。シフトモード別。CAP表(THM設備Cap/機種対応表)の機種別キャパ列から生成。
# 機種×設備の生産可否(○/△/×)の帰結を織り込んだ「その機種がラインで出せる日次台数」。
# Suica4 は全設備×(生産不可)のため載せていない — 計画側でキャパ未定義=生産不可として除外する。
PRODUCT_DAILY_CAPS_BY_MODE: dict[str, dict[str, float]] = {
    "22h": {
        "さそり金融": 120000,
        "SuicaⅢ": 120000,
        "さそり交通": 120000,
        "Lite-S(Mies)": 42240,
        "部分リライト": 30000,
        "SD3": 30000,
        "SD-T1": 30000,
        "MOT2": 60000,
    },
    "16h": {
        "さそり金融": 80000,
        "SuicaⅢ": 80000,
        "さそり交通": 80000,
        "Lite-S(Mies)": 30720,
        "部分リライト": 20000,
        "SD3": 20000,
        "SD-T1": 20000,
        "MOT2": 20000,
    },
    "11h": {
        "さそり金融": 60000,
        "SuicaⅢ": 60000,
        "さそり交通": 60000,
        "Lite-S(Mies)": 21120,
        "部分リライト": 15000,
        "SD3": 15000,
        "SD-T1": 15000,
        "MOT2": 15000,
    },
    "8h": {
        "さそり金融": 40000,
        "SuicaⅢ": 40000,
        "さそり交通": 40000,
        "Lite-S(Mies)": 15360,
        "部分リライト": 10000,
        "SD3": 10000,
        "SD-T1": 10000,
        "MOT2": 10000,
    },
}

# 完成品コード(RC-…の基底)→ 機種呼称。機種一覧(CAP)から生成。
PRODUCT_ALIASES: dict[str, str] = {
    "RC-S100": "さそり金融",
    "RC-SA02F": "さそり金融",
    "RC-S103": "さそり交通",
    "RC-S104": "さそり交通",
    "RC-SA05A": "さそり交通",
    "RC-S105": "SuicaⅢ",
    "RC-S106": "SuicaⅢ",
    "RC-SA06A": "SuicaⅢ",
    "RC-S982F": "Lite-S(Mies)",
    "RC-SA10A": "部分リライト",
    "RC-S123": "SD-T1",
    "RC-SA15A": "SD-T1",
    "RC-S125": "Suica4",
    "RC-S127": "MOT2",
    "RC-S140": "SD3",
    "RC-SA42F": "SD3",
}


@dataclass
class UnmappedRow:
    row: int
    order_id: str
    name: str


def _norm(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s　]", "", str(value)).upper()


def resolve_product(name: object, aliases: dict[str, str] | None = None) -> str | None:
    """完成品名/コードから機種呼称を最長一致で解決する。該当なしは None。"""
    aliases = aliases or PRODUCT_ALIASES
    n = _norm(name)
    best: str | None = None
    for key in aliases:
        if n.startswith(key) and (best is None or len(key) > len(best)):
            best = key
    return aliases[best] if best else None


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _header_index(ws, header_row: int) -> dict[str, int]:
    idx: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        if name is not None:
            idx[str(name).strip()] = c
    return idx


def parse_thm_ledger(
    file_obj: BinaryIO,
    aliases: dict[str, str] | None = None,
    sheet_name: str = "台帳",
    header_row: int = 2,
    only_due_on_or_after: date | None = None,
    lines: set[str] | None = None,
    shipment_buffer_days: int = 2,
) -> tuple[list[DemandItem], list[UnmappedRow]]:
    """台帳ワークブックを需要(DemandItem)一覧へ変換する。

    **納期の捉え方**: 本当の納期は台帳の**出荷日**。完成はその `shipment_buffer_days`(既定2)
    日前に済ませたいので、EDD・遅れ判定に使う `due_date`(完成目標日)=出荷日−バッファ(暦日)
    とする。出荷日が無い行のみ**完成予定日**へフォールバック(この場合バッファは掛けない)。
    `ship_date` には出荷日(表示用の真の納期)を入れる。

    - `only_due_on_or_after`: 完成目標日(=出荷日−バッファ)がこの日以降の受注だけ対象にする。
    - `lines`: ライン名の集合を渡すと、その受注だけに絞る(例: {"CTA1", "CTA2"})。
    戻り値: (需要一覧, 呼称解決できなかった行一覧)。
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    idx = _header_index(ws, header_row)

    def col(*names: str) -> int | None:
        for n in names:
            if n in idx:
                return idx[n]
        return None

    c_no = col("№", "No", "No.")
    c_line = col("ライン")
    c_name = col("完成品名")
    c_code = col("完成品コード")
    c_seiban = col("製番")
    c_qty = col("完成予定数")
    c_comp = col("完成予定日")
    c_ship = col("出荷日")

    demands: list[DemandItem] = []
    unmapped: list[UnmappedRow] = []
    buffer = timedelta(days=shipment_buffer_days)

    for r in range(header_row + 1, ws.max_row + 1):
        row_no = ws.cell(row=r, column=c_no).value if c_no else None
        if not row_no:
            continue
        row_no = str(row_no).strip()
        # 出荷ロットの識別子は製番列(現場のMIL表と同じキー)。製番が空の行のみ№で代用。
        seiban = ws.cell(row=r, column=c_seiban).value if c_seiban else None
        order_id = str(seiban).strip() if seiban not in (None, "", "-") else row_no

        if lines and c_line:
            line = ws.cell(row=r, column=c_line).value
            if str(line).strip() not in lines:
                continue

        qty = ws.cell(row=r, column=c_qty).value if c_qty else None
        ship = _as_date(ws.cell(row=r, column=c_ship).value) if c_ship else None
        comp = _as_date(ws.cell(row=r, column=c_comp).value) if c_comp else None
        # 完成目標日 = 出荷日 − バッファ。出荷日が無ければ完成予定日にフォールバック。
        if ship is not None:
            due = ship - buffer
        else:
            due = comp
        if not isinstance(qty, (int, float)) or qty <= 0 or due is None:
            continue
        if only_due_on_or_after and due < only_due_on_or_after:
            continue

        name = ws.cell(row=r, column=c_name).value if c_name else None
        code = ws.cell(row=r, column=c_code).value if c_code else None
        product = resolve_product(name, aliases) or resolve_product(code, aliases)
        if product is None:
            unmapped.append(UnmappedRow(row=r, order_id=order_id, name=str(name)))
            continue

        demands.append(
            DemandItem(product=product, quantity=float(qty), due_date=due, order_id=order_id, ship_date=ship)
        )

    return demands, unmapped


STOPS_SHEET_NAMES = ("01_設備停止マスタ", "設備停止")


def parse_equipment_stops(file_obj: BinaryIO, header_row: int = 1) -> list["EquipmentStop"]:
    """設備停止マスタ(01_設備停止マスタ / 設備停止 シート)を読み取る。

    有効=Y の行だけを対象にする(現場テンプレートの判定ロジックどおり)。
    シートが無ければ空リストを返す。
    """
    from bottleneck_planner import EquipmentStop

    wb = load_workbook(file_obj, data_only=True)
    ws = None
    for name in STOPS_SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return []

    idx = _header_index(ws, header_row)

    def val(row: int, *names: str):
        for n in names:
            if n in idx:
                return ws.cell(row=row, column=idx[n]).value
        return None

    def num(v) -> float | None:
        return float(v) if isinstance(v, (int, float)) else None

    stops: list[EquipmentStop] = []
    for r in range(header_row + 1, ws.max_row + 1):
        stop_id = val(r, "停止ID")
        if not stop_id:
            continue
        if str(val(r, "有効") or "").strip().upper() != "Y":
            continue
        start_day = _as_date(val(r, "開始日"))
        end_day = _as_date(val(r, "終了日"))
        stage_id = str(val(r, "工程") or "").strip()
        if start_day is None or end_day is None or not stage_id:
            continue
        stops.append(
            EquipmentStop(
                stop_id=str(stop_id).strip(),
                stage_id=stage_id,
                machine_id=str(val(r, "設備") or "").strip(),
                start_day=start_day,
                end_day=end_day,
                start_shift=str(val(r, "開始勤務") or "A勤").strip(),
                end_shift=str(val(r, "終了勤務") or "B勤").strip(),
                method=str(val(r, "停止Cap控除方法") or "全停止").strip(),
                stop_rate_pct=num(val(r, "停止率_%", "停止率")),
                stop_hours=num(val(r, "停止時間_h", "停止時間")),
                corrected_cap=num(val(r, "補正後Cap_台", "補正後Cap")),
                reason=str(val(r, "停止区分") or "").strip(),
            )
        )
    return stops


DAILY_ACTUALS_SHEET_NAMES = ("日次実績", "実績_日次", "進捗実績")


def parse_daily_actuals(file_obj: BinaryIO, header_row: int = 1) -> dict[date, float]:
    """「日次実績」シート(日付 / [工程] / 実績数)から、日付別の生産実績(ライン計)を読む。

    工程列があっても、進捗はライン計で見るため日付ごとに合算する(製番別総数の
    `parse_actuals` とは別物; こちらは日々の進捗管理に使う)。シート無しは空dict。
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = None
    for name in DAILY_ACTUALS_SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return {}

    idx = _header_index(ws, header_row)
    c_day = idx.get("日付") or idx.get("Date")
    c_qty = next((idx[h] for h in ("実績数", "実績数量", "実績", "台数") if h in idx), None)
    if c_day is None or c_qty is None:
        return {}

    result: dict[date, float] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        d = _as_date(ws.cell(row=r, column=c_day).value)
        qty = ws.cell(row=r, column=c_qty).value
        if d is None or not isinstance(qty, (int, float)):
            continue
        result[d] = result.get(d, 0.0) + float(qty)
    return result


ACTUALS_SHEET_NAMES = ("実績", "実績反映")
_ACTUALS_QTY_HEADERS = ("実績数", "実績数量", "実績")


def parse_actuals(file_obj: BinaryIO, header_row: int = 1) -> dict[str, float]:
    """「実績」シート(列: 製番 / 実績数)から、製番別の生産実績数量を読み取る。

    シートが無ければ空dictを返す(実績なし=そのまま立案)。同じ製番が複数行ある場合は合算。
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = None
    for name in ACTUALS_SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return {}

    idx = _header_index(ws, header_row)
    c_seiban = idx.get("製番")
    c_qty = next((idx[h] for h in _ACTUALS_QTY_HEADERS if h in idx), None)
    if c_seiban is None or c_qty is None:
        return {}

    actuals: dict[str, float] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        seiban = ws.cell(row=r, column=c_seiban).value
        qty = ws.cell(row=r, column=c_qty).value
        if seiban in (None, "") or not isinstance(qty, (int, float)) or qty <= 0:
            continue
        key = str(seiban).strip()
        actuals[key] = actuals.get(key, 0.0) + float(qty)
    return actuals


def _is_red_font(cell) -> bool:
    """セルの文字色が赤(FFFF0000)か。現場の表では赤字=実績、黒字=予定。"""
    fc = cell.font.color
    return bool(fc and isinstance(fc.rgb, str) and fc.rgb == "FFFF0000")


def parse_thm_shortterm_actuals(file_obj: BinaryIO, sheet_name: str = "TA1") -> dict[str, dict[str, float]]:
    """THM短期投入予定表から、製番別に TAL(投入)/MIL(完成) の**実績(赤字)**を合算する。

    レイアウト: 製番ごとに「Line-In行(=TAL投入, ピンク)」と直後の「Completion行(=MIL完成,
    オレンジ)」の2行ペア。列6="Line-In"、列3=製番、列7以降が日付列(列1-6はラベル/累計)。
    各行の日付列のうち**赤字**セル(=実績)だけを合算する。黒字(予定)は無視。
    戻り値: {製番: {"TAL": 実績台数, "MIL": 実績台数}}(実績が1つでもある製番のみ)。
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    first_date_col = 7
    out: dict[str, dict[str, float]] = {}

    def row_red_sum(r: int) -> float:
        total = 0.0
        for c in range(first_date_col, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, (int, float)) and _is_red_font(cell):
                total += float(v)
        return total

    r = 5
    while r <= ws.max_row:
        role = ws.cell(row=r, column=6).value
        seiban = ws.cell(row=r, column=3).value
        if role == "Line-In" and seiban not in (None, "", "-"):
            key = str(seiban).strip()
            tal = row_red_sum(r)
            mil = row_red_sum(r + 1)
            if tal or mil:
                cur = out.setdefault(key, {"TAL": 0.0, "MIL": 0.0})
                cur["TAL"] += tal
                cur["MIL"] += mil
            r += 2
        else:
            r += 1
    return out


def parse_ta1_hal_actuals(
    file_obj: BinaryIO, year: int, sheet_name: str = "生産計画"
) -> dict[date, float]:
    """TA1_投入計画(生産計画シート)から、HAL工程の**実績(赤字)**を日別に合算する。

    レイアウト: 機種ごとにブロック(TAL/TAL累計/HAL/HAL累計/MIL...)。列1が工程ラベル。
    行1に月ラベル(例 "9月", 列持ち越し)、行2に日番号。列3以降が日付。HAL行(列1=="HAL")の
    日付セルのうち**赤字**を全機種ブロック横断で日別合算する。`year` は開始年(月が戻ったら+1)。
    戻り値: {date: HAL実績台数}。
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    col_date: dict[int, date] = {}
    cur_month: int | None = None
    cur_year = year
    for c in range(3, ws.max_column + 1):
        mv = ws.cell(row=1, column=c).value
        if isinstance(mv, str) and mv.endswith("月"):
            try:
                m = int(mv[:-1])
                if cur_month is not None and m < cur_month:
                    cur_year += 1
                cur_month = m
            except ValueError:
                pass
        dv = ws.cell(row=2, column=c).value
        if cur_month and isinstance(dv, (int, float)):
            try:
                col_date[c] = date(cur_year, cur_month, int(dv))
            except ValueError:
                pass

    hal: dict[date, float] = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value != "HAL":
            continue
        for c, d in col_date.items():
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, (int, float)) and _is_red_font(cell):
                hal[d] = hal.get(d, 0.0) + float(v)
    return hal
