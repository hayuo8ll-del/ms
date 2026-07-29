"""生産計画の自動立案 vs 手動計画(FeliCa) 照合機構そのものの検証。

照合の答え(MAE)が信用できるかを、既知の正解を持つ合成データで確かめる。
"""
import io
import sys
from datetime import date
from pathlib import Path

BACKEND = Path(__file__).resolve().parent.parent / "backend"
sys.path.insert(0, str(BACKEND))

from openpyxl import Workbook

from bottleneck_planner import (
    DemandItem, StageFlowConfig, plan_bottleneck, working_days_in_range,
)
from felica_calibration import (
    calibrate, compare_plans, parse_felica_plan, _nearest_index, _with_offsets,
)

CAPS = {"16h": 90000, "22h": 120000}


def felica_workbook(rows, sheet="202607_CTA1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    dates = sorted({d for _s, _p, _l, li, co in rows for d in list(li) + list(co)})
    for c, label in enumerate(
        ["Planner", "WS", "Item Desc", "Plan", "Item No", "Before", "Lot", "Date"], 1
    ):
        ws.cell(row=3, column=c, value=label)
    date_col = {d: 9 + i for i, d in enumerate(dates)}
    for d, c in date_col.items():
        ws.cell(row=3, column=c, value=d)
    r = 5
    for seiban, product, lot, li, co in rows:
        ws.cell(row=r, column=3, value=product)
        ws.cell(row=r, column=4, value=seiban)
        ws.cell(row=r, column=7, value=lot)
        ws.cell(row=r, column=8, value="Line-In")
        ws.cell(row=r + 1, column=8, value="Completion")
        for d, q in li.items():
            ws.cell(row=r, column=date_col[d], value=q)
        for d, q in co.items():
            ws.cell(row=r + 1, column=date_col[d], value=q)
        r += 2
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def hdr(n, t):
    print(f"\n{'='*72}\n[{n}] {t}\n{'='*72}")


# ---------------------------------------------------------------- 1
hdr(1, "_with_offsets が leadOffsetByProduct(機種別オフセット)を落とす")
flows = [
    StageFlowConfig("ANT", -1),
    StageFlowConfig("TAL", -2, input_unit=40000),
    StageFlowConfig("HAL", 0, input_unit=10000),
    StageFlowConfig("MIL", 2, input_unit=1920,
                    lead_offset_by_product={"さそり金融": 5}),
]
print("入力 MIL.lead_offset_by_product =", flows[3].lead_offset_by_product)
out = _with_offsets(flows, {"ANT": -1, "TAL": -2, "HAL": 0, "MIL": 2})
print("_with_offsets 通過後      =", out[3].lead_offset_by_product)
print("→ 同じオフセットを渡しても機種別上書きが消える:",
      "消える(BUG)" if out[3].lead_offset_by_product is None else "保持")
print("offset_for('さそり金融'): 前 =", flows[3].offset_for("さそり金融"),
      "/ 後 =", out[3].offset_for("さそり金融"))

# ---------------------------------------------------------------- 2
hdr(2, "calibrate() の current(現行config)が機種別オフセットを無視して測られる")
days = working_days_in_range(date(2026, 7, 1), date(2026, 7, 31))
caps = {"16h": {"さそり金融": 60000}, "22h": {"さそり金融": 90000}}
demands = [DemandItem("さそり金融", 120000, date(2026, 7, 31), order_id="S1")]

base = plan_bottleneck(demands, days, CAPS, stage_flows=flows,
                       product_caps_by_mode=caps, a_shift_only_switch=False)
mil_done = max(c.day for c in base.stage_allocation if c.stage_id == "MIL")
print("機種別オフセット(+5)を効かせた計画の MIL完成日 =", mil_done)

stripped = plan_bottleneck(demands, days, CAPS, stage_flows=out,
                           product_caps_by_mode=caps, a_shift_only_switch=False)
mil_done2 = max(c.day for c in stripped.stage_allocation if c.stage_id == "MIL")
print("calibrate内部(_with_offsets後)の MIL完成日   =", mil_done2)
print("→ 較正が評価している『現行config』は実際の計画と別物:",
      mil_done != mil_done2)

# ---------------------------------------------------------------- 3
hdr(3, "_nearest_index が計画期間外の日付を端の稼働日へ丸める(誤差の過小評価)")
wd_index = {d: i for i, d in enumerate(days)}
for probe in [date(2026, 4, 1), date(2026, 6, 30), date(2026, 12, 25)]:
    i = _nearest_index(probe, wd_index, days)
    print(f"  FeliCa側の日付 {probe} → 稼働日index {i} ({days[i]}) "
          f"実際は {(probe - days[0]).days:+d} 暦日ずれ")
print("→ 期間外ロットも『照合できた』扱いになり、差は端で頭打ちになる。")

buf = felica_workbook([("S1", "RC-SA02F/5", 120000,
                        {date(2026, 4, 1): 120000}, {date(2026, 4, 30): 120000})])
fel = parse_felica_plan(buf)
rep = compare_plans(base, fel, days)
print(f"  4月完成のFeliCaロットを7月計画と照合 → matched={rep.matched}, "
      f"completion_mae={rep.completion_mae} 稼働日")
print("  (実際は約3か月ずれているのに、MAEは計画期間の幅で頭打ち)")

# ---------------------------------------------------------------- 4
hdr(4, "同日carryover除外が、連続流動ロットの日次を大量に削る")
li = {date(2026, 7, d): 10000 for d in (1, 2, 3, 6, 7, 8, 9, 10)}
co = {date(2026, 7, d): 10000 for d in (3, 6, 7, 8, 9, 10, 13, 14)}
buf = felica_workbook([("S9", "RC-SA02F/5", 80000, li, co)])
fel = parse_felica_plan(buf)
lot = fel["S9"]
print(f"元データ: Line-In {len(li)}日 / Completion {len(co)}日 (リードタイム2稼働日の通常ロット)")
print(f"除外後  : Line-In {len(lot.line_in_daily)}日 / Completion {len(lot.completion_daily)}日")
print("  残った Line-In   :", sorted(lot.line_in_daily))
print("  残った Completion:", sorted(lot.completion_daily))
print(f"  line_in_first={lot.line_in_first} / completion_last={lot.completion_last}")
print("→ 『投入と完成が同日=リードタイム0』の前提は、"
      "ロット内で流動が重なる期間には当てはまらない。")

# ---------------------------------------------------------------- 5
hdr(5, "照合の往復テスト: FeliCa = 自分の計画そのもの なら誤差0になるか")
mil = {}
ant = {}
for c in base.stage_allocation:
    if c.stage_id == "MIL":
        mil[c.day] = mil.get(c.day, 0) + c.quantity
    elif c.stage_id == "ANT":
        ant[c.day] = ant.get(c.day, 0) + c.quantity
buf = felica_workbook([("S1", "RC-SA02F/5", 120000, ant, mil)])
fel = parse_felica_plan(buf)
rep = compare_plans(base, fel, days, aliases={"RC-SA02F": "さそり金融"})
print(f"自計画をそのままFeliCa形式にして照合:")
print(f"  matched={rep.matched} completion_mae={rep.completion_mae} "
      f"start_mae={rep.start_mae}")
print(f"  日次形状 completion_daily_mae={rep.completion_daily_mae} "
      f"line_in_daily_mae={rep.line_in_daily_mae}")
print("  期待値: すべて0。0でなければ照合機構自体にバイアスがある。")

# ---------------------------------------------------------------- 6
hdr(6, "Line-In ラベルの表記ゆれで照合結果が静かに0件になる")
wb = Workbook()
ws = wb.active
ws.title = "202607_CTA1"
ws.cell(row=3, column=9, value=date(2026, 7, 1))
ws.cell(row=5, column=3, value="RC-SA02F/5")
ws.cell(row=5, column=4, value="S1")
ws.cell(row=5, column=8, value="Line-In ")   # 末尾に空白
ws.cell(row=6, column=8, value="Completion")
ws.cell(row=5, column=9, value=90000)
b = io.BytesIO()
wb.save(b)
b.seek(0)
print("col8='Line-In '(末尾空白) のとき 読めた製番 =", list(parse_felica_plan(b)))
print("→ 製番(col4)は .strip() しているが、役割ラベル(col8)は完全一致比較。"
      "表記ゆれ・全角・改行で照合が黙って0件になる。")

# ---------------------------------------------------------------- 7
hdr(7, "matched が完成日ペアのみを数える(投入日の母数が見えない)")
buf = felica_workbook([
    ("S1", "RC-SA02F/5", 120000, {date(2026, 7, 1): 120000}, {}),   # 完成なし
])
fel = parse_felica_plan(buf)
rep = compare_plans(base, fel, days)
print(f"完成日が無く投入日だけ一致するロット → matched={rep.matched} "
      f"(start_mae={rep.start_mae} は算出済み)")
print("→ 画面の『照合できた製番数』は完成日の母数。投入日MAEの母数は表示されない。")
