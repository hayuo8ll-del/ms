"""実データ4点で「自動立案 vs 手計画」を計測する。

  python3 docs/verify_real_data.py --ledger 台帳.xlsx --felica FeliCa.xlsx \
      --thm THM短期.xlsx --ta1 TA1.xlsx

計測するもの:
  1. 取り込み(台帳/THM短期/TA1/FeliCa)が正しく読めているか
  2. HAL日次配分を、あなたのTA1_投入計画(黒字=予定, 赤字=実績)と直接突き合わせ
  3. A勤限定切替 ON/OFF での納期遅れ・遊休能力・並行機種数
  4. 計画期間を変えたときの FeliCa 照合MAE の振れ
  5. 祝日(FeliCa灰色セル)の反映有無による工程別の欠損
"""
from __future__ import annotations

import argparse
import io
import sys
import warnings
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))
warnings.filterwarnings("ignore")

from openpyxl import load_workbook  # noqa: E402

from bottleneck_planner import (  # noqa: E402
    A_SHIFT_DEFERRAL_TAG,
    apply_actuals,
    plan_bottleneck,
    working_days_in_range,
)
from config_loader import load_bottleneck_planning  # noqa: E402
from felica_calibration import (  # noqa: E402
    compare_plans,
    parse_felica_nonworking_days,
    parse_felica_plan,
)
from thm_ledger_import import (  # noqa: E402
    parse_ta1_hal_actuals,
    parse_thm_ledger,
    parse_thm_shortterm_actuals,
)

# TA1_投入計画の機種コード -> 呼称(config の productAliases は RC- 付きなので別途持つ)
TA1_CODE_TO_NAME = {
    "A02F": "さそり金融", "S100": "さそり金融",
    "A05A": "さそり交通", "S103": "さそり交通",
    "S105": "SuicaⅢ", "982F": "Lite-S(Mies)", "S127": "MOT2",
    "A10A": "部分リライト", "A42F": "SD3", "A15A": "SD-T1",
}
RED = ("FF0000", "FFFF0000", "00FF0000")
# 機種別ブロックは行88まで。行90-92は全機種合計(数式)なので機種別集計から外す。
TA1_PRODUCT_BLOCK_END = 89


def _is_red(cell) -> bool:
    f = cell.font
    if f is None or f.color is None:
        return False
    c = f.color
    if c.type == "rgb" and c.rgb:
        return str(c.rgb).upper() in RED
    if c.type == "indexed":
        return c.indexed in (2, 10)
    return False


def parse_ta1_plan_and_actual(path: str, year: int = 2026):
    """TA1_投入計画を 機種×工程×日 で読む。黒字=予定(人が立てた計画), 赤字=実績。"""
    wbv, wbs = load_workbook(path, data_only=True), load_workbook(path)
    wsv, wss = wbv["生産計画"], wbs["生産計画"]

    col_date: dict[int, date] = {}
    month = None
    for c in range(3, wsv.max_column + 1):
        m = wsv.cell(row=1, column=c).value
        if isinstance(m, str) and m.endswith("月"):
            month = int(m[:-1])
        elif isinstance(m, (int, float)):
            month = int(m)
        d = wsv.cell(row=2, column=c).value
        if month and isinstance(d, (int, float)):
            try:
                col_date[c] = date(year, month, int(d))
            except ValueError:
                pass

    plan: dict[str, dict[str, dict[date, float]]] = {}
    actual: dict[str, dict[str, dict[date, float]]] = {}
    product = None
    for r in range(1, TA1_PRODUCT_BLOCK_END):
        label = wsv.cell(row=r, column=1).value
        code = wsv.cell(row=r, column=2).value
        if isinstance(code, str) and code.strip():
            product = code.strip()
        if not isinstance(label, str) or label.strip() not in ("TAL", "HAL", "MIL", "ANT"):
            continue
        stage = label.strip()
        if product is None:
            continue
        for c, d in col_date.items():
            v = wsv.cell(row=r, column=c).value
            if not isinstance(v, (int, float)) or not v:
                continue
            tgt = actual if _is_red(wss.cell(row=r, column=c)) else plan
            tgt.setdefault(product, {}).setdefault(stage, {})
            tgt[product][stage][d] = tgt[product][stage].get(d, 0.0) + float(v)
    return plan, actual


def ta1_stage_by_day(src, stage, lo, hi) -> dict[date, dict[str, float]]:
    out: dict[date, dict[str, float]] = {}
    for code, stages in src.items():
        name = TA1_CODE_TO_NAME.get(code)
        if not name:
            continue
        for d, q in stages.get(stage, {}).items():
            if lo <= d <= hi:
                out.setdefault(d, {})
                out[d][name] = out[d].get(name, 0.0) + q
    return out


def hal_by_day(result) -> dict[date, dict[str, float]]:
    out: dict[date, dict[str, float]] = {}
    for c in result.stage_allocation:
        if c.stage_id == "HAL":
            out.setdefault(c.day, {})
            out[c.day][c.product] = out[c.day].get(c.product, 0.0) + c.quantity
    return out


def changeovers(by_day: dict[date, dict[str, float]]) -> int:
    n, prev = 0, set()
    for d in sorted(by_day):
        cur = set(by_day[d])
        n += len(cur - prev)
        prev = cur
    return n


def hdr(t: str) -> None:
    print(f"\n{'=' * 78}\n{t}\n{'=' * 78}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ledger", required=True, help="THM生産台帳")
    ap.add_argument("--felica", required=True, help="FeliCa実計画")
    ap.add_argument("--thm", help="THM短期投入予定表(赤字=実績)")
    ap.add_argument("--ta1", help="TA1_投入計画(黒字=予定/赤字=実績)")
    ap.add_argument("--start", default="2026-07-29", help="計画開始日")
    ap.add_argument("--end", default="2026-08-31", help="計画終了日")
    args = ap.parse_args()
    start, end = date.fromisoformat(args.start), date.fromisoformat(args.end)

    cfg = load_bottleneck_planning()
    ledger_bytes = Path(args.ledger).read_bytes()

    # ---------------------------------------------------------------- 1
    hdr("1. 取り込みの健全性")
    demands, unmapped = parse_thm_ledger(
        io.BytesIO(ledger_bytes), aliases=cfg.product_aliases,
        shipment_buffer_days=cfg.shipment_buffer_days,
    )
    print(f"  台帳       : {len(demands)}製番 / 機種未解決 {len(unmapped)}件 / "
          f"総需要 {sum(d.quantity for d in demands):,.0f}台")
    due = [d.due_date for d in demands]
    print(f"               完成目標 {min(due)} 〜 {max(due)}")

    actuals: dict[str, float] = {}
    if args.thm:
        thm = parse_thm_shortterm_actuals(open(args.thm, "rb"))
        actuals = {s: v["MIL"] for s, v in thm.items() if v.get("MIL")}
        tal = sum(v.get("TAL", 0.0) for v in thm.values())
        print(f"  THM短期    : 赤字MIL実績 {len(actuals)}製番 / "
              f"{sum(actuals.values()):,.0f}台  (TAL実績 {tal:,.0f}台)")
    if args.ta1:
        hal_act = parse_ta1_hal_actuals(open(args.ta1, "rb"), year=start.year)
        print(f"  TA1        : 赤字HAL実績 {len(hal_act)}日 / {sum(hal_act.values()):,.0f}台")

    felica = parse_felica_plan(open(args.felica, "rb"))
    live = {k: v for k, v in felica.items() if v.line_in_daily or v.completion_daily}
    holidays = parse_felica_nonworking_days(open(args.felica, "rb"))
    print(f"  FeliCa     : {len(felica)}製番行 (台数のある行 {len(live)}) / "
          f"台帳と共通の製番 {len({d.order_id for d in demands} & set(felica))}件")
    print(f"  FeliCa灰色 : 非稼働日 {[str(d) for d in holidays]}")
    print(f"  config     : nonWorkingDays = {cfg.non_working_days or '(空)'}"
          f"{'   ← 祝日が未反映' if not cfg.non_working_days and holidays else ''}")

    full_demands = list(demands)  # 実績控除前(FeliCa照合はこちら: /validate と同条件)
    if actuals:
        demands, _w = apply_actuals(demands, actuals)
    remaining = sum(d.quantity for d in demands)

    def build(a_switch=True, hol=None, s=start, e=end, dem=None):
        wd = working_days_in_range(s, e, holidays=hol or None)
        return wd, plan_bottleneck(
            dem if dem is not None else demands, wd, cfg.line_daily_capacities,
            stage_flows=cfg.stage_flows,
            a_shift_only_switch=a_switch, a_shift_fraction=cfg.a_shift_fraction,
            product_caps_by_mode=cfg.product_daily_caps_by_mode,
            bottleneck_stage=cfg.bottleneck_stage, machine_counts=cfg.machine_counts,
        )

    # ---------------------------------------------------------------- 2
    if args.ta1:
        hdr(f"2. HAL投入 {start}〜{end}: あなたのTA1計画 vs 自動立案")
        ta1_plan, ta1_actual = parse_ta1_plan_and_actual(args.ta1, year=start.year)
        yours = ta1_stage_by_day(ta1_plan, "HAL", start, end)
        _wd, result = build()
        ours = hal_by_day(result)
        print(f"  自動立案: モード {result.shift_mode} / 能力 {result.daily_capacity:,.0f}台日 / "
              f"残需要 {remaining:,.0f}台")
        print(f"\n  {'日':<12}{'あなた':>9}{'自動':>9}{'差':>10}   自動の構成 / あなたの構成")
        tot = 0.0
        days = sorted(set(ours) | set(yours))
        for d in days:
            a = sum(yours.get(d, {}).values())
            b = sum(ours.get(d, {}).values())
            tot += abs(b - a)
            f = lambda m: "+".join(f"{p[:5]}{q/1000:.0f}k" for p, q in sorted(m.items()) if q)
            print(f"  {str(d):<12}{a:>9,.0f}{b:>9,.0f}{b - a:>+10,.0f}   "
                  f"{f(ours.get(d, {})):<32} / {f(yours.get(d, {}))}")
        print(f"\n  日別ライン計の平均絶対差 = {tot / len(days):,.0f} 台/日")

        ay = ta1_stage_by_day(ta1_actual, "HAL", date(start.year, 1, 1), end)
        for lbl, m in (("あなたの実績(TA1赤字)", ay), ("自動立案", ours)):
            if not m:
                continue
            print(f"  {lbl:<22}: {len(m)}稼働日 / 平均並行機種数 "
                  f"{sum(len(v) for v in m.values()) / len(m):.2f} / "
                  f"複数機種の日 {sum(1 for v in m.values() if len(v) > 1)}/{len(m)} / "
                  f"立上げ+切替 {changeovers(m)}回")

    # ---------------------------------------------------------------- 3
    hdr("3. A勤限定切替 ON/OFF の影響")
    for lbl, sw in (("ON(現行)", True), ("OFF", False)):
        _wd, r = build(a_switch=sw)
        h = hal_by_day(r)
        idle = sum(max(0.0, r.daily_capacity - sum(v.values())) for v in h.values())
        print(f"  A勤限定切替 {lbl:<8} 納期遅れ {sum(1 for l in r.mil_lots if not l.on_time):>3}製番 / "
              f"遊休能力 {idle:>10,.0f}台 / 平均並行機種数 "
              f"{sum(len(v) for v in h.values()) / len(h):.2f} / "
              f"繰下げ警告 {sum(1 for w in r.warnings if A_SHIFT_DEFERRAL_TAG in w):>3}件")
        print(f"      先頭3日: " + "  ".join(
            f"{d} {{{', '.join(f'{p}:{q/1000:.0f}k' for p, q in v.items())}}}"
            for d, v in sorted(h.items())[:3]))

    # ---------------------------------------------------------------- 4
    hdr("4. 計画期間を変えたときの FeliCa 照合MAE の振れ")
    print("  (FeliCaは月全体の計画なので、照合は実績控除前の需要で立案した計画に対して行う)")
    print(f"  {'期間':<24}{'稼働日':>6}{'モード':>7}{'照合':>6}{'完成MAE':>9}"
          f"{'完成bias':>10}{'投入MAE':>9}")
    for s, e in ((date(2026, 7, 1), date(2026, 7, 31)),
                 (date(2026, 7, 1), date(2026, 8, 15)),
                 (date(2026, 7, 1), date(2026, 8, 31)),
                 (date(2026, 7, 1), date(2026, 9, 30))):
        wd, r = build(s=s, e=e, dem=full_demands)
        rep = compare_plans(r, felica, wd, aliases=cfg.product_aliases)
        print(f"  {f'{s}..{e}':<24}{len(wd):>6}{r.shift_mode:>7}{rep.matched:>6}"
              f"{rep.completion_mae:>9}{rep.completion_bias:>+10}{rep.start_mae:>9}")

    wd, r = build(s=date(2026, 7, 1), e=date(2026, 8, 31), dem=full_demands)
    rep = compare_plans(r, felica, wd, aliases=cfg.product_aliases)
    print(f"\n  機種別タイミング差 (our − FeliCa; 正=自動立案が遅い)")
    print(f"  {'機種':<16}{'完成bias':>10}{'完成MAE':>9}{'投入bias':>10}{'n':>4}")
    for p, v in sorted(rep.timing_by_product.items(),
                       key=lambda kv: -abs(kv[1]["completion_bias"])):
        print(f"  {p:<16}{v['completion_bias']:>+10}{v['completion_mae']:>9}"
              f"{v['start_bias']:>+10}{v['n']:>4}")

    # ---------------------------------------------------------------- 5
    hdr("5. 祝日(FeliCa灰色セル)の反映有無による工程別の欠損")
    for lbl, hol in (("祝日なし(現行config)", None), ("FeliCaの祝日を反映", set(holidays))):
        wd, r = build(hol=hol)
        st: dict[str, float] = {}
        for c in r.stage_allocation:
            st[c.stage_id] = st.get(c.stage_id, 0.0) + c.quantity
        print(f"\n  --- {lbl} --- 稼働日{len(wd)} モード{r.shift_mode} "
              f"遅れ{sum(1 for l in r.mil_lots if not l.on_time)}製番")
        for s in ("ANT", "TAL", "HAL", "MIL"):
            q = st.get(s, 0.0)
            print(f"      {s}: {q:>11,.0f}  需要比 {q / remaining * 100:5.1f}%  "
                  f"欠損 {remaining - q:>10,.0f}")


if __name__ == "__main__":
    main()
